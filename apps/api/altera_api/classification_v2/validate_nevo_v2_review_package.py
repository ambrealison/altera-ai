"""Phase Quality-V2-S — read-only validator for a FILLED NEVO V2 review package.

A reviewer fills the ``manual_decision`` (and, for replacements, the
``approved_*``) columns of the dry-run review package
(``nevo_v2_enrich_review_package_<project>.csv``, or a filtered bucket CSV).
This tool validates those decisions OFFLINE and writes a small report so the
team can see what is apply-ready, what is blocked, and what still needs work.

It is strictly read-only: it reads ONE CSV/XLSX file and writes report
artifacts. It never touches the database, never imports a route, never
activates V2, and is not a runtime dependency of the app.

    python -m altera_api.classification_v2.validate_nevo_v2_review_package \
        --input /tmp/altera-quality/nevo_v2_enrich_review_package_<id>.csv \
        --output-dir /tmp/altera-quality

CSV is the primary, always-supported input. XLSX is supported only when
``openpyxl`` is installed; if an ``.xlsx`` is given without it, the tool fails
with a clear message (openpyxl is NOT a required runtime dependency).
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

from altera_api.classification_v2.nevo_review_workflow import (
    MANUAL_DECISION_VALUES,
)

# blank counts as "pending"; everything else must be a known verb.
ALLOWED_DECISIONS = frozenset(MANUAL_DECISION_VALUES) | {""}

#: sheets in an xlsx package that are NOT row data.
_META_SHEETS = frozenset({"Summary", "Instructions"})

_OVERRIDE_MARKER = "OVERRIDE"

_PROJECT_RE = re.compile(
    r"nevo_v2_enrich_review_package_(?P<pid>.+)\.(?:csv|xlsx)$"
)

RECOMMENDATIONS = (
    "ready_for_apply_planning", "review_incomplete", "blocked_by_errors",
)

ERROR_CSV_COLUMNS = [
    "product_id", "product_name", "manual_decision", "review_priority",
    "suggested_action", "nutrition_safety_action", "message",
]
APPROVED_CSV_COLUMNS = [
    "product_id", "product_name", "manual_decision", "source",
    "effective_nevo_code", "effective_nevo_name",
    "effective_protein_g_per_100g", "review_priority", "suggested_action",
    "reviewer_notes",
]


class ReviewPackageError(Exception):
    """Raised for an unreadable / malformed review package."""


def _s(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_number(text: str) -> bool:
    try:
        float(text)
    except (TypeError, ValueError):
        return False
    return True


def infer_project_id(input_path: str | Path) -> str | None:
    m = _PROJECT_RE.search(Path(input_path).name)
    return m.group("pid") if m else None


# ---------------------------------------------------------------------------
# Input reading — CSV primary, XLSX optional.
# ---------------------------------------------------------------------------
def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on runtime env
        raise ReviewPackageError(
            f"input {path.name} is an .xlsx file but openpyxl is not installed. "
            "Install openpyxl, or export/validate the CSV review package "
            "(nevo_v2_enrich_review_package_<project>.csv) instead."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in wb.worksheets:
        if sheet.title in _META_SHEETS:
            continue
        it = sheet.iter_rows(values_only=True)
        try:
            header = [_s(c) for c in next(it)]
        except StopIteration:
            continue
        if "manual_decision" not in header:
            continue
        for values in it:
            if values is None or all(v is None for v in values):
                continue
            rows.append({h: values[i] if i < len(values) else None
                         for i, h in enumerate(header)})
    return rows


def read_review_package(input_path: str | Path) -> list[dict[str, Any]]:
    path = Path(input_path)
    if not path.exists():
        raise ReviewPackageError(f"input file not found: {path}")
    if path.suffix.lower() == ".xlsx":
        rows = _read_xlsx(path)
    elif path.suffix.lower() == ".csv":
        rows = _read_csv(path)
    else:
        raise ReviewPackageError(
            f"unsupported input extension {path.suffix!r} (use .csv or .xlsx)"
        )
    if not rows:
        raise ReviewPackageError(f"no data rows found in {path.name}")
    if "manual_decision" not in rows[0]:
        raise ReviewPackageError(
            f"{path.name} is missing a 'manual_decision' column — is this a "
            "filled review package?"
        )
    return rows


# ---------------------------------------------------------------------------
# Per-row validation (Part A + Part B).
# ---------------------------------------------------------------------------
def validate_row(row: dict[str, Any]) -> tuple[list[str], list[str]]:
    """Return ``(errors, warnings)`` for one filled review row."""
    errors: list[str] = []
    warnings: list[str] = []

    decision = _s(row.get("manual_decision")).lower()
    if decision not in ALLOWED_DECISIONS:
        errors.append(
            f"invalid manual_decision {decision!r} "
            "(allowed: approve, reject, replace, needs_more_info, or blank)"
        )
        return errors, warnings

    priority = _s(row.get("review_priority")).upper()
    suggested = _s(row.get("suggested_action"))
    nsa = _s(row.get("nutrition_safety_action"))
    notes_upper = _s(row.get("reviewer_notes")).upper()
    has_override = _OVERRIDE_MARKER in notes_upper

    appr_code = _s(row.get("approved_nevo_code"))
    appr_name = _s(row.get("approved_nevo_name"))
    appr_protein = _s(row.get("approved_protein_g_per_100g"))
    ex_code = _s(row.get("nevo_code"))
    ex_name = _s(row.get("nevo_food_name"))
    ex_protein = _s(row.get("enriched_protein_g_per_100g"))

    if appr_protein and not _is_number(appr_protein):
        errors.append(
            f"approved_protein_g_per_100g must be numeric, got {appr_protein!r}"
        )

    is_non_food = suggested in ("reject_non_food", "reject_policy_excluded")

    if decision == "replace":
        if not appr_code or not appr_name:
            errors.append(
                "replace requires approved_nevo_code and approved_nevo_name"
            )
    elif decision == "approve":
        if priority == "P0":
            errors.append("P0 row cannot be approved (high-risk / never auto)")
        if is_non_food:
            if has_override:
                warnings.append(
                    "approved a non-food / policy-excluded row via OVERRIDE — "
                    "confirm this is really a food product"
                )
            else:
                errors.append(
                    "non-food / policy-excluded row cannot be approved without "
                    f"an explicit {_OVERRIDE_MARKER} marker in reviewer_notes"
                )

        effective_code = appr_code or ex_code
        if not effective_code:
            errors.append(
                "approve on a no_match row requires approved_nevo_code"
            )
        elif appr_code:
            if not appr_name:
                errors.append(
                    "approve with approved_nevo_code also requires "
                    "approved_nevo_name"
                )
        else:  # approving the existing candidate as-is
            missing = [
                name for name, val in (
                    ("nevo_code", ex_code), ("nevo_food_name", ex_name),
                    ("enriched_protein_g_per_100g", ex_protein),
                ) if not val
            ]
            if missing:
                errors.append(
                    "approve of the existing candidate requires "
                    + ", ".join(missing)
                )
            # Risk-aware warnings for approving a downgraded row as-is.
            if nsa == "skip_state_mismatch":
                warnings.append(
                    "approved a state-mismatch row as-is — the reference "
                    "physical state differs (verify nutrition basis)"
                )
            elif nsa == "skip_proxy_too_broad":
                warnings.append(
                    "approved a proxy-too-broad row as-is — the reference is a "
                    "processing/variety proxy (verify)"
                )
            elif nsa == "route_to_review" or suggested == "review_generic_proxy":
                warnings.append(
                    "approved a route-to-review / generic-proxy row as-is "
                    "(verify before enriching)"
                )
    # reject / needs_more_info / blank: no approved fields required.
    return errors, warnings


def _effective_source(row: dict[str, Any]) -> dict[str, str]:
    appr_code = _s(row.get("approved_nevo_code"))
    if appr_code:
        return {
            "source": "replacement",
            "effective_nevo_code": appr_code,
            "effective_nevo_name": _s(row.get("approved_nevo_name")),
            "effective_protein_g_per_100g": _s(
                row.get("approved_protein_g_per_100g")
            ),
        }
    return {
        "source": "existing",
        "effective_nevo_code": _s(row.get("nevo_code")),
        "effective_nevo_name": _s(row.get("nevo_food_name")),
        "effective_protein_g_per_100g": _s(row.get("enriched_protein_g_per_100g")),
    }


# ---------------------------------------------------------------------------
# Whole-package validation + summary.
# ---------------------------------------------------------------------------
def validate_package(
    rows: list[dict[str, Any]], *, input_path: str, project_id: str | None,
) -> dict[str, Any]:
    error_records: list[dict[str, Any]] = []
    warning_records: list[dict[str, Any]] = []
    approved_records: list[dict[str, Any]] = []

    counts = {
        "approve": 0, "reject": 0, "replace": 0, "needs_more_info": 0,
        "pending": 0,
    }
    blocked_count = 0
    apply_ready_count = 0

    for row in rows:
        decision = _s(row.get("manual_decision")).lower()
        bucket = "pending" if decision == "" else decision
        counts[bucket] = counts.get(bucket, 0) + 1

        errors, warnings = validate_row(row)
        meta = {
            "product_id": _s(row.get("product_id")),
            "product_name": _s(row.get("product_name")),
            "manual_decision": decision,
            "review_priority": _s(row.get("review_priority")),
            "suggested_action": _s(row.get("suggested_action")),
            "nutrition_safety_action": _s(row.get("nutrition_safety_action")),
        }
        for msg in errors:
            error_records.append({**meta, "message": msg})
        for msg in warnings:
            warning_records.append({**meta, "message": msg})

        if errors:
            blocked_count += 1
        elif decision in ("approve", "replace"):
            apply_ready_count += 1
            approved_records.append({
                "product_id": meta["product_id"],
                "product_name": meta["product_name"],
                "manual_decision": decision,
                **_effective_source(row),
                "review_priority": meta["review_priority"],
                "suggested_action": meta["suggested_action"],
                "reviewer_notes": _s(row.get("reviewer_notes")),
            })

    error_count = len(error_records)
    if error_count:
        recommendation = "blocked_by_errors"
    elif counts["pending"] or counts["needs_more_info"]:
        recommendation = "review_incomplete"
    else:
        recommendation = "ready_for_apply_planning"

    summary = {
        "input_path": input_path,
        "project_id": project_id,
        "total_rows": len(rows),
        "approved_count": counts["approve"],
        "rejected_count": counts["reject"],
        "replace_count": counts["replace"],
        "needs_more_info_count": counts["needs_more_info"],
        "pending_count": counts["pending"],
        "error_count": error_count,
        "warning_count": len(warning_records),
        "apply_ready_count": apply_ready_count,
        "blocked_count": blocked_count,
        "recommendation": recommendation,
    }
    return {
        "summary": summary,
        "errors": error_records,
        "warnings": warning_records,
        "approved": approved_records,
    }


def _write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in columns})


def write_artifacts(
    out_dir: str | Path, project_id: str, result: dict[str, Any],
) -> dict[str, str]:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    json_path = out / f"nevo_v2_review_validation_summary_{project_id}.json"
    errors_path = out / f"nevo_v2_review_validation_errors_{project_id}.csv"
    warnings_path = out / f"nevo_v2_review_validation_warnings_{project_id}.csv"
    approved_path = out / f"nevo_v2_review_approved_candidates_{project_id}.csv"

    json_path.write_text(
        json.dumps(result["summary"], indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    _write_csv(errors_path, ERROR_CSV_COLUMNS, result["errors"])
    _write_csv(warnings_path, ERROR_CSV_COLUMNS, result["warnings"])
    _write_csv(approved_path, APPROVED_CSV_COLUMNS, result["approved"])
    return {
        "summary_json": str(json_path),
        "errors_csv": str(errors_path),
        "warnings_csv": str(warnings_path),
        "approved_candidates_csv": str(approved_path),
    }


def build_arg_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        prog="python -m altera_api.classification_v2."
             "validate_nevo_v2_review_package",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--input", required=True,
                    help="filled review package (.csv primary, .xlsx optional)")
    ap.add_argument("--output-dir", default="/tmp/altera-quality")
    ap.add_argument(
        "--project-id", default=None,
        help="override the project id inferred from the filename (use this when "
             "validating a renamed/copied sample package).",
    )
    return ap


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    try:
        rows = read_review_package(args.input)
    except ReviewPackageError as exc:
        print(f"FATAL: {exc}")
        return 2

    resolved_project_id = args.project_id or infer_project_id(args.input)
    project_id = resolved_project_id or "unknown"
    result = validate_package(
        rows, input_path=str(args.input), project_id=resolved_project_id,
    )
    paths = write_artifacts(args.output_dir, project_id, result)

    s = result["summary"]
    print("# NEVO V2 review-package validation (READ-ONLY — no database writes)")
    print(f"  input={s['input_path']} project={s['project_id'] or 'n/a'} "
          f"rows={s['total_rows']}")
    print("-" * 64)
    print(f"  approve={s['approved_count']}  reject={s['rejected_count']}  "
          f"replace={s['replace_count']}  "
          f"needs_more_info={s['needs_more_info_count']}  "
          f"pending={s['pending_count']}")
    print(f"  errors={s['error_count']}  warnings={s['warning_count']}  "
          f"apply_ready={s['apply_ready_count']}  blocked={s['blocked_count']}")
    print(f"  RECOMMENDATION: {s['recommendation']}")
    print("-" * 64)
    for label, key in (
        ("Summary JSON", "summary_json"), ("Errors CSV", "errors_csv"),
        ("Warnings CSV", "warnings_csv"),
        ("Approved candidates CSV", "approved_candidates_csv"),
    ):
        print(f"  {label}: {paths[key]}")
    print("READ-ONLY — no database writes were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

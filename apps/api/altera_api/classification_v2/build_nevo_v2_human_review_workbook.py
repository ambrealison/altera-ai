"""Phase Quality-V2-AG — human-friendly NEVO V2 review workbook (read-only).

Turns the machine review package (Quality-V2-AF) into something a non-engineer
can actually use: a multi-tab Excel workbook (if openpyxl is available) with the
useful columns first, technical columns tucked away, priority colouring, and
decision dropdowns — plus a clean CSV fallback and a plain-text README when
openpyxl is missing (e.g. on Render). The validator stays the source of truth;
this only reshapes the package for humans.

    python -m altera_api.classification_v2.build_nevo_v2_human_review_workbook \
        --project-id <uuid> --output-dir /tmp/altera-quality
        # --review-package <path> overrides auto-discovery of the newest
        # nevo_v2_batch_review_package_<project>_*.csv

No DB writes. No apply plan. V1 stays default; embeddings stay off. No routes.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

from altera_api.classification_v2.apply_nevo_v2_plan import _s

try:  # openpyxl is optional — Render images may not ship it.
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - exercised via _HAS_OPENPYXL flag
    _HAS_OPENPYXL = False

# --- Human-facing column order (Part C) ------------------------------------
HUMAN_COLUMNS = [
    "review_priority", "review_source", "product_name", "category", "brand",
    "safety_action", "suggested_action", "current_batch_nevo_code",
    "current_batch_nevo_name", "current_batch_protein_g_per_100g",
    "existing_v2_nevo_code", "existing_v2_nevo_name", "top_5_candidate_names",
    "top_5_candidate_codes", "rejection_summary", "manual_decision",
    "reviewer_notes", "approved_nevo_code", "approved_nevo_name",
    "approved_protein_g_per_100g", "alias_candidate", "rule_candidate",
    "gold_case_decision",
]
#: Technical columns are appended after the human block so the file is
#: self-contained (round-trips through the normalizer) but the reviewer sees
#: the useful columns first.
TECHNICAL_TAIL = [
    "project_id", "run_id", "product_id", "canonical_product_key",
    "ingredients", "duplicate_count", "v2_outcome", "confidence", "match_type",
    "top_5_similarities", "diff_bucket", "batch_matches_existing_v2",
]
HUMAN_FILE_COLUMNS = [*HUMAN_COLUMNS, *TECHNICAL_TAIL]

#: Editable reviewer columns (highlighted / unlocked in the workbook).
EDITABLE_COLUMNS = [
    "manual_decision", "reviewer_notes", "approved_nevo_code",
    "approved_nevo_name", "approved_protein_g_per_100g", "alias_candidate",
    "rule_candidate", "gold_case_decision",
]
LONG_TEXT_COLUMNS = {
    "product_name", "ingredients", "top_5_candidate_names",
    "top_5_candidate_codes", "rejection_summary", "reviewer_notes",
}

MANUAL_DECISIONS = [
    "approve_existing_candidate", "approve_existing_v2", "replace", "reject",
    "needs_more_info", "out_of_scope",
]
GOLD_DECISIONS = [
    "positive_gold", "negative_gold", "alias_candidate", "rule_candidate",
    "ignore",
]

# Sort orders (Part D).
_PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
_SOURCE_ORDER = {
    "existing_v2_diff": 0, "safety_downgrade": 1, "needs_review": 2,
    "no_match": 3,
}
_PRIORITY_FILL = {
    "P0": "F4CCCC", "P1": "FCE5CD", "P2": "FFF2CC", "P3": "EFEFEF",
}

# Per-bucket human sheets (Part B).
_BUCKET_SHEETS = [
    ("Safety_Downgrade", "safety_downgrade"),
    ("Needs_Review", "needs_review"),
    ("No_Match", "no_match"),
    ("Existing_V2_Diffs", "existing_v2_diff"),
]


def _to_human_row(pkg: dict[str, Any]) -> dict[str, str]:
    """Map a raw package row to the human file row (friendly batch names)."""
    row: dict[str, str] = {}
    for col in HUMAN_FILE_COLUMNS:
        if col == "current_batch_nevo_code":
            row[col] = _s(pkg.get("batch_nevo_code")) or _s(pkg.get("nevo_code"))
        elif col == "current_batch_nevo_name":
            row[col] = (_s(pkg.get("batch_nevo_name"))
                        or _s(pkg.get("nevo_food_name")))
        elif col == "current_batch_protein_g_per_100g":
            row[col] = _s(pkg.get("protein_g_per_100g"))
        else:
            row[col] = _s(pkg.get(col))
    return row


def _sort_key(row: dict[str, str]) -> tuple[int, int, str]:
    return (
        _PRIORITY_ORDER.get(row.get("review_priority", ""), 9),
        _SOURCE_ORDER.get(row.get("review_source", ""), 9),
        row.get("product_name", "").lower(),
    )


def human_rows(pkg_rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    return sorted((_to_human_row(r) for r in pkg_rows), key=_sort_key)


# --- README / Instructions text --------------------------------------------
def _instruction_lines(project_id: str, run_id: str, total: int) -> list[str]:
    return [
        "NEVO V2 — human review workbook (READ-ONLY pipeline, no database)",
        "=================================================================",
        f"project_id : {project_id}",
        f"run_id     : {run_id}",
        f"rows       : {total}",
        "",
        "WHAT THIS IS",
        "------------",
        "These are products our automatic NEVO V2 matcher could NOT safely",
        "enrich on its own. A human needs to confirm or correct the match.",
        "Nothing here is live: filling this file changes no production data.",
        "",
        "HOW TO REVIEW",
        "-------------",
        "1. Start on the 'P1_Review_First' tab (highest priority).",
        "2. For each row, read product_name / category / brand and the",
        "   suggested match (current_batch_nevo_*) or candidates",
        "   (top_5_candidate_names).",
        "3. Put your decision in 'manual_decision' (dropdown).",
        "4. Add 'reviewer_notes' when needed (especially overrides).",
        "5. For 'replace', also fill approved_nevo_code + approved_nevo_name",
        "   (and approved_protein_g_per_100g if you know it).",
        "6. Leave a row blank to mark it as still pending.",
        "",
        "manual_decision values",
        "-----------------------",
        "approve_existing_candidate : accept the current batch NEVO candidate",
        "approve_existing_v2        : keep the existing V2 match",
        "replace                    : you give approved_nevo_code +"
        " approved_nevo_name",
        "reject                     : no NEVO match should be used",
        "needs_more_info            : you cannot decide yet",
        "out_of_scope               : product should not be enriched at all",
        "(blank)                    : pending / not yet reviewed",
        "",
        "SAFETY OVERRIDES",
        "----------------",
        "Safety_Downgrade rows were blocked for a reason. To approve the batch",
        "candidate anyway, add the matching token to reviewer_notes:",
        "  OVERRIDE_SAFE_STATE  -> for a state mismatch (e.g. raw vs cooked)",
        "  OVERRIDE_SAFE_PROXY  -> for a proxy that is too broad",
        "P0 rows (if any) require the word OVERRIDE in reviewer_notes to"
        " approve.",
        "",
        "gold_case_decision values (optional, for building training data)",
        "---------------------------------------------------------------",
        "positive_gold  : a confirmed correct match worth keeping as gold",
        "negative_gold  : a confirmed wrong match worth keeping as a negative",
        "alias_candidate: product name suggests a reusable alias",
        "rule_candidate : suggests a reusable matching rule",
        "ignore         : not useful as a training case",
        "",
        "WHEN DONE",
        "---------",
        "Save the file, then hand it back. We normalize it and run the",
        "validator (the source of truth):",
        "  python -m altera_api.classification_v2."
        "normalize_nevo_v2_human_review_workbook \\",
        f"      --input <your_filled_file> --project-id {project_id} \\",
        "      --output-dir /tmp/altera-quality",
        "  python -m altera_api.classification_v2."
        "validate_nevo_v2_batch_review_package \\",
        "      --input <the_FILLED_NORMALIZED_csv> "
        f"--project-id {project_id} \\",
        "      --output-dir /tmp/altera-quality",
        "",
        "The 'Technical_Raw' tab / trailing columns are for engineers — you do",
        "not need to touch them.",
    ]


_REFERENCE_ROWS = [
    ("field", "value", "meaning"),
    *[("manual_decision", v, m) for v, m in [
        ("approve_existing_candidate", "accept the current batch NEVO"
         " candidate"),
        ("approve_existing_v2", "keep the existing V2 match"),
        ("replace", "reviewer provides approved_nevo_code +"
         " approved_nevo_name"),
        ("reject", "no NEVO match should be used"),
        ("needs_more_info", "reviewer cannot decide yet"),
        ("out_of_scope", "product should not be enriched"),
        ("(blank)", "pending / not yet reviewed"),
    ]],
    ("reviewer_notes", "OVERRIDE_SAFE_STATE", "approve batch despite state"
     " mismatch"),
    ("reviewer_notes", "OVERRIDE_SAFE_PROXY", "approve batch despite too-broad"
     " proxy"),
    ("reviewer_notes", "OVERRIDE", "required to approve a P0 row"),
    *[("gold_case_decision", v, m) for v, m in [
        ("positive_gold", "confirmed correct match (gold)"),
        ("negative_gold", "confirmed wrong match (negative gold)"),
        ("alias_candidate", "name suggests a reusable alias"),
        ("rule_candidate", "suggests a reusable matching rule"),
        ("ignore", "not useful as a training case"),
    ]],
]


# --- CSV fallback -----------------------------------------------------------
def _write_csv(path: Path, columns: list[str], rows: list[dict[str, str]],
               ) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


# --- XLSX workbook ----------------------------------------------------------
def _style_sheet(ws: Any, columns: list[str], n_rows: int) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    edit_fill = PatternFill("solid", fgColor="E2EFDA")
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = edit_fill if col in EDITABLE_COLUMNS else header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        letter = get_column_letter(idx)
        if col in LONG_TEXT_COLUMNS:
            width = 42
        elif col in ("manual_decision", "gold_case_decision",
                     "current_batch_nevo_name", "existing_v2_nevo_name"):
            width = 26
        else:
            width = 18
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{max(n_rows + 1, 1)}"
    # Wrap long-text body cells.
    long_idx = [i for i, c in enumerate(columns, start=1)
                if c in LONG_TEXT_COLUMNS]
    for r in range(2, n_rows + 2):
        for i in long_idx:
            ws.cell(row=r, column=i).alignment = Alignment(
                vertical="top", wrap_text=True)


def _color_priorities(ws: Any, columns: list[str], rows: list[dict[str, str]],
                      ) -> None:
    if "review_priority" not in columns:
        return
    for r, row in enumerate(rows, start=2):
        fill_hex = _PRIORITY_FILL.get(row.get("review_priority", ""))
        if not fill_hex:
            continue
        fill = PatternFill("solid", fgColor=fill_hex)
        for c in range(1, len(columns) + 1):
            # Keep editable columns visually distinct (light green).
            if columns[c - 1] in EDITABLE_COLUMNS:
                continue
            ws.cell(row=r, column=c).fill = fill


def _add_dropdowns(ws: Any, columns: list[str], n_rows: int) -> None:
    specs = [
        ("manual_decision", MANUAL_DECISIONS),
        ("gold_case_decision", GOLD_DECISIONS),
    ]
    last = max(n_rows + 1, 2)
    for col_name, options in specs:
        if col_name not in columns:
            continue
        letter = get_column_letter(columns.index(col_name) + 1)
        dv = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"',
            allow_blank=True, showDropDown=False)
        dv.error = "Pick a value from the list (or leave blank for pending)."
        dv.prompt = "Choose: " + ", ".join(options)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last}")


def _fill_sheet(ws: Any, columns: list[str], rows: list[dict[str, str]], *,
                dropdowns: bool) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    _style_sheet(ws, columns, len(rows))
    _color_priorities(ws, columns, rows)
    if dropdowns:
        _add_dropdowns(ws, columns, len(rows))


def _build_workbook(rows: list[dict[str, str]], *, project_id: str,
                    run_id: str, raw_rows: list[dict[str, Any]],
                    raw_columns: list[str]) -> Any:
    wb = openpyxl.Workbook()

    ws_inst = wb.active
    ws_inst.title = "Instructions"
    for line in _instruction_lines(project_id, run_id, len(rows)):
        ws_inst.append([line])
    ws_inst.column_dimensions["A"].width = 78

    _fill_sheet(wb.create_sheet("Review_All"), HUMAN_FILE_COLUMNS, rows,
                dropdowns=True)

    p1_rows = [r for r in rows if r.get("review_priority") in ("P0", "P1")]
    _fill_sheet(wb.create_sheet("P1_Review_First"), HUMAN_COLUMNS, p1_rows,
                dropdowns=True)

    for title, source in _BUCKET_SHEETS:
        bucket_rows = [r for r in rows if r.get("review_source") == source]
        _fill_sheet(wb.create_sheet(title), HUMAN_COLUMNS, bucket_rows,
                    dropdowns=True)

    ws_ref = wb.create_sheet("Reference_Decisions")
    for ref in _REFERENCE_ROWS:
        ws_ref.append(list(ref))
    for letter, width in (("A", 18), ("B", 28), ("C", 52)):
        ws_ref.column_dimensions[letter].width = width
    for c in range(1, 4):
        ws_ref.cell(row=1, column=c).font = Font(bold=True)
    ws_ref.freeze_panes = "A2"

    ws_raw = wb.create_sheet("Technical_Raw")
    ws_raw.append(raw_columns)
    for raw in raw_rows:
        ws_raw.append([_s(raw.get(col)) for col in raw_columns])
    ws_raw.freeze_panes = "A2"
    return wb


# --- auto-discovery ---------------------------------------------------------
def _auto_discover(out_dir: Path, project_id: str) -> Path | None:
    pattern = f"nevo_v2_batch_review_package_{project_id}_*.csv"
    matches = sorted(out_dir.glob(pattern), key=lambda p: p.stat().st_mtime)
    return matches[-1] if matches else None


def _run_id_from(path: Path, project_id: str) -> str:
    m = re.search(rf"_{re.escape(project_id)}_(.+)\.csv$", path.name)
    return m.group(1) if m else "run"


def _read_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


# --- orchestration ----------------------------------------------------------
def build(*, project_id: str, review_package: Path, output_dir: Path,
          run_id: str | None = None) -> dict[str, Any]:
    pkg_rows = _read_rows(review_package)
    raw_columns = (list(pkg_rows[0].keys()) if pkg_rows
                   else list(REVIEW_PACKAGE_FALLBACK))
    rows = human_rows(pkg_rows)
    run = run_id or _run_id_from(review_package, project_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    base = f"{project_id}_{run}"
    paths: dict[str, str] = {}

    # CSV fallback is ALWAYS written (works with or without openpyxl).
    csv_path = output_dir / f"nevo_v2_human_review_workbook_{base}.csv"
    _write_csv(csv_path, HUMAN_FILE_COLUMNS, rows)
    paths["csv_fallback"] = str(csv_path)

    readme_path = output_dir / f"nevo_v2_human_review_README_{base}.txt"
    readme_path.write_text(
        "\n".join(_instruction_lines(project_id, run, len(rows))) + "\n",
        encoding="utf-8")
    paths["readme"] = str(readme_path)

    xlsx_written = False
    if _HAS_OPENPYXL:
        wb = _build_workbook(rows, project_id=project_id, run_id=run,
                             raw_rows=pkg_rows, raw_columns=raw_columns)
        xlsx_path = output_dir / f"nevo_v2_human_review_workbook_{base}.xlsx"
        wb.save(xlsx_path)
        paths["workbook_xlsx"] = str(xlsx_path)
        xlsx_written = True

    by_priority = {p: sum(1 for r in rows if r.get("review_priority") == p)
                   for p in ("P0", "P1", "P2", "P3")}
    by_source = {s: sum(1 for r in rows if r.get("review_source") == s)
                 for s in {r.get("review_source", "") for r in rows} if s}
    summary = {
        "project_id": project_id, "run_id": run,
        "total_rows": len(rows),
        "count_by_priority": by_priority,
        "count_by_review_source": by_source,
        "xlsx_written": xlsx_written,
        "openpyxl_available": _HAS_OPENPYXL,
        "review_package_input": str(review_package),
        "human_columns": HUMAN_COLUMNS,
        "technical_columns": TECHNICAL_TAIL,
        "manual_decision_values": MANUAL_DECISIONS,
        "gold_case_decision_values": GOLD_DECISIONS,
        "output_paths": paths,
    }
    summary_path = output_dir / f"nevo_v2_human_review_summary_{base}.json"
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False),
                            encoding="utf-8")
    paths["summary"] = str(summary_path)
    summary["output_paths"] = paths
    return summary


#: Column fallback if the review package is empty (header-only file).
REVIEW_PACKAGE_FALLBACK = HUMAN_FILE_COLUMNS


def build_arg_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        prog="python -m altera_api.classification_v2."
             "build_nevo_v2_human_review_workbook",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--project-id", required=True)
    ap.add_argument("--review-package", default=None)
    ap.add_argument("--output-dir", default="/tmp/altera-quality")
    ap.add_argument("--run-id", default=None)
    return ap


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    out_dir = Path(args.output_dir)
    pid = _s(args.project_id)

    if args.review_package:
        pkg = Path(args.review_package)
    else:
        pkg = _auto_discover(out_dir, pid)
        if pkg is None:
            print("ERROR: no review package found. Build one first with "
                  "build_nevo_v2_batch_review_package, or pass "
                  "--review-package.")
            return 2
    if not pkg.exists():
        print(f"ERROR: review package not found: {pkg}")
        return 2

    summary = build(project_id=pid, review_package=pkg, output_dir=out_dir,
                    run_id=args.run_id)

    print("# NEVO V2 human review workbook (read-only — no database writes)")
    print(f"  project={pid} run_id={summary['run_id']} "
          f"rows={summary['total_rows']}")
    print(f"  by_priority={summary['count_by_priority']}")
    print(f"  by_source={summary['count_by_review_source']}")
    if summary["xlsx_written"]:
        print(f"  Workbook XLSX: {summary['output_paths']['workbook_xlsx']}")
    else:
        print("  openpyxl NOT available — wrote CSV fallback + README "
              "(export to Excel locally if you want a workbook).")
    print(f"  CSV fallback:  {summary['output_paths']['csv_fallback']}")
    print(f"  README:        {summary['output_paths']['readme']}")
    print(f"  Summary JSON:  {summary['output_paths']['summary']}")
    print("READ-ONLY — no database writes were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

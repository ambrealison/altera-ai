"""Phase Quality-V2-AG — normalize a filled human workbook back to validator input.

The human workbook/CSV (Quality-V2-AG) renames a few columns (current_batch_*)
and reorders them for readability. This CLI turns a FILLED human file (xlsx or
csv) back into the canonical batch review package CSV so the existing validator
runs unchanged:

    python -m altera_api.classification_v2.normalize_nevo_v2_human_review_workbook \
        --input /path/to/filled_review.xlsx_or_csv \
        --output-dir /tmp/altera-quality --project-id <uuid>

    python -m altera_api.classification_v2.validate_nevo_v2_batch_review_package \
        --input .../nevo_v2_batch_review_package_FILLED_NORMALIZED_<p>_<run>.csv \
        --output-dir /tmp/altera-quality --project-id <uuid>

XLSX needs openpyxl. If the input is .xlsx and openpyxl is missing, this fails
with a clear instruction to export the sheet as CSV from Excel. No DB writes.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from altera_api.classification_v2.apply_nevo_v2_plan import _s
from altera_api.classification_v2.build_nevo_v2_batch_review_package import (
    REVIEW_PACKAGE_COLUMNS,
)

try:
    import openpyxl

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - exercised via flag
    _HAS_OPENPYXL = False

#: The sheet the reviewer edits (self-contained: human + technical columns).
_PRIMARY_SHEET = "Review_All"
#: Sheets we never treat as the data source.
_NON_DATA_SHEETS = {
    "Instructions", "Reference_Decisions", "Technical_Raw", "P1_Review_First",
    "Safety_Downgrade", "Needs_Review", "No_Match", "Existing_V2_Diffs",
}


class NormalizeError(RuntimeError):
    """A human-facing normalization failure."""


def _to_package_row(human: dict[str, Any]) -> dict[str, str]:
    """Map a human file row back to the canonical package row."""
    def g(*keys: str) -> str:
        for k in keys:
            v = _s(human.get(k))
            if v:
                return v
        return ""

    batch_code = g("current_batch_nevo_code", "batch_nevo_code", "nevo_code")
    batch_name = g("current_batch_nevo_name", "batch_nevo_name",
                   "nevo_food_name")
    protein = g("current_batch_protein_g_per_100g", "protein_g_per_100g")

    row = {col: g(col) for col in REVIEW_PACKAGE_COLUMNS}
    row["nevo_code"] = batch_code
    row["nevo_food_name"] = batch_name
    row["batch_nevo_code"] = batch_code
    row["batch_nevo_name"] = batch_name
    row["protein_g_per_100g"] = protein
    return row


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if not _HAS_OPENPYXL:
        raise NormalizeError(
            "Input is .xlsx but openpyxl is not installed here. Open the file "
            "in Excel, choose the 'Review_All' sheet, and 'Save As' / export "
            "it as CSV, then re-run this command on the .csv.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = _pick_sheet(wb.sheetnames)
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    columns = [str(c) if c is not None else "" for c in header]
    out: list[dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(v in (None, "") for v in values):
            continue
        record = {columns[i]: ("" if v is None else v)
                  for i, v in enumerate(values) if i < len(columns)}
        out.append(record)
    wb.close()
    return out


def _pick_sheet(sheetnames: list[str]) -> str:
    if _PRIMARY_SHEET in sheetnames:
        return _PRIMARY_SHEET
    for name in sheetnames:
        if name not in _NON_DATA_SHEETS:
            return name
    # Fall back to the first data-ish sheet even if reserved.
    return sheetnames[0]


def read_input_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    if suffix in (".csv", ".txt", ""):
        return _read_csv_rows(path)
    raise NormalizeError(
        f"Unsupported input type '{suffix}'. Provide a .csv or .xlsx file.")


def normalize(*, input_path: Path, output_dir: Path, project_id: str | None,
              run_id: str | None) -> dict[str, Any]:
    if not input_path.exists():
        raise NormalizeError(f"input not found: {input_path}")
    human = read_input_rows(input_path)
    rows = [_to_package_row(r) for r in human]

    pid = _s(project_id) or (rows[0].get("project_id") if rows else "") or "p"
    run = run_id or (rows[0].get("run_id") if rows else "") or "run"
    if project_id:  # explicit override wins for every row.
        for r in rows:
            r["project_id"] = pid
    if run_id:
        for r in rows:
            r["run_id"] = run

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = (output_dir
                / f"nevo_v2_batch_review_package_FILLED_NORMALIZED_"
                  f"{pid}_{run}.csv")
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=REVIEW_PACKAGE_COLUMNS,
                                extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    filled = sum(1 for r in rows if _s(r.get("manual_decision")))
    return {
        "project_id": pid, "run_id": run, "rows": len(rows),
        "rows_with_decision": filled, "output_path": str(out_path),
        "input_path": str(input_path),
    }


def build_arg_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(
        prog="python -m altera_api.classification_v2."
             "normalize_nevo_v2_human_review_workbook",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--input", required=True)
    ap.add_argument("--output-dir", default="/tmp/altera-quality")
    ap.add_argument("--project-id", default=None)
    ap.add_argument("--run-id", default=None)
    return ap


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    try:
        result = normalize(
            input_path=Path(args.input), output_dir=Path(args.output_dir),
            project_id=args.project_id, run_id=args.run_id)
    except NormalizeError as exc:
        print(f"ERROR: {exc}")
        return 2

    print("# NEVO V2 human review normalize (read-only — no database writes)")
    print(f"  project={result['project_id']} run_id={result['run_id']} "
          f"rows={result['rows']} with_decision={result['rows_with_decision']}")
    print(f"  Normalized package CSV: {result['output_path']}")
    print("  Next: run validate_nevo_v2_batch_review_package on that CSV.")
    print("READ-ONLY — no database writes were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

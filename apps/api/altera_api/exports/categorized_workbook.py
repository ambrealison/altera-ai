"""Retailer-facing categorised export workbook (.xlsx).

Builds an Excel workbook a retailer can download once their catalogue is
categorised:

* **Sheet 1 — Produits / Products**: every product with its Protein Tracker
  and WWF categories + source + confidence (non-commercial fields only).
* **Sheet 2 — Analyse Protein Tracker**: count per PT group with a bar chart
  + a plant-vs-animal pie chart.
* **Sheet 3 — Analyse WWF**: count per WWF food group with a bar chart.

Pure + deterministic: it takes already-computed rows and returns the xlsx
bytes. No store / network access here, so it is trivially unit-testable.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Row model
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExportRow:
    """One product's categorised, non-commercial export record."""

    external_product_id: str
    product_name: str
    retailer_category: str | None
    pt_group: str | None
    pt_source: str | None
    pt_confidence: float | None
    wwf_food_group: str | None
    wwf_subgroup: str | None
    wwf_composite_bucket: str | None
    wwf_source: str | None
    wwf_confidence: float | None
    # Per-product Protein Tracker amounts (kg), from the latest PT run.
    # None when no calculation has run yet for this product.
    pt_plant_protein_kg: float | None = None
    pt_animal_protein_kg: float | None = None
    pt_total_protein_kg: float | None = None


# ---------------------------------------------------------------------------
# Localised labels (the workbook content is localised via a ``lang`` arg —
# it is file content, not app UI, so it is not part of the web i18n dicts).
# ---------------------------------------------------------------------------

_L = {
    "fr": {
        "sheet_products": "Produits",
        "sheet_pt": "Analyse Protein Tracker",
        "sheet_wwf": "Analyse WWF",
        "h_id": "Réf. produit",
        "h_name": "Produit",
        "h_retailer": "Catégorie distributeur",
        "h_pt_group": "Protein Tracker",
        "h_pt_source": "Source PT",
        "h_pt_conf": "Confiance PT",
        "h_pt_plant_kg": "Protéines végétales (kg)",
        "h_pt_animal_kg": "Protéines animales (kg)",
        "h_pt_plant_share": "Part végétale (%)",
        "h_wwf_group": "Groupe WWF",
        "h_wwf_sub": "Sous-groupe WWF",
        "h_wwf_comp": "Composite (bucket)",
        "h_wwf_source": "Source WWF",
        "h_wwf_conf": "Confiance WWF",
        "a_category": "Catégorie",
        "a_count": "Nombre de produits",
        "a_kg": "Protéines (kg)",
        "a_pt_title": "Répartition Protein Tracker",
        "a_pt_chart": "Produits par groupe PT",
        "a_pt_pie": "Végétal vs animal (nombre de produits)",
        "a_pt_protein_title": "Répartition des protéines (kg)",
        "a_pt_protein_pie": "Protéines végétales vs animales (kg)",
        "a_wwf_title": "Répartition WWF",
        "a_wwf_chart": "Produits par groupe alimentaire WWF",
        "plant": "Végétal",
        "animal": "Animal",
        "composite": "Composite",
        "other": "Autre",
        "title": "Export catégorisé",
    },
    "en": {
        "sheet_products": "Products",
        "sheet_pt": "Protein Tracker analysis",
        "sheet_wwf": "WWF analysis",
        "h_id": "Product ref.",
        "h_name": "Product",
        "h_retailer": "Retailer category",
        "h_pt_group": "Protein Tracker",
        "h_pt_source": "PT source",
        "h_pt_conf": "PT confidence",
        "h_pt_plant_kg": "Plant protein (kg)",
        "h_pt_animal_kg": "Animal protein (kg)",
        "h_pt_plant_share": "Plant share (%)",
        "h_wwf_group": "WWF group",
        "h_wwf_sub": "WWF subgroup",
        "h_wwf_comp": "Composite (bucket)",
        "h_wwf_source": "WWF source",
        "h_wwf_conf": "WWF confidence",
        "a_category": "Category",
        "a_count": "Product count",
        "a_kg": "Protein (kg)",
        "a_pt_title": "Protein Tracker distribution",
        "a_pt_chart": "Products per PT group",
        "a_pt_pie": "Plant vs animal (product count)",
        "a_pt_protein_title": "Protein split (kg)",
        "a_pt_protein_pie": "Plant vs animal protein (kg)",
        "a_wwf_title": "WWF distribution",
        "a_wwf_chart": "Products per WWF food group",
        "plant": "Plant",
        "animal": "Animal",
        "composite": "Composite",
        "other": "Other",
        "title": "Categorised export",
    },
}

_PT_GROUP_LABELS = {
    "fr": {
        "plant_based_core": "Végétal — cœur",
        "plant_based_non_core": "Végétal — hors cœur",
        "composite_products": "Composite",
        "animal_core": "Animal — cœur",
        "out_of_scope": "Hors périmètre",
        "unknown": "Inconnu",
    },
    "en": {
        "plant_based_core": "Plant — core",
        "plant_based_non_core": "Plant — non-core",
        "composite_products": "Composite",
        "animal_core": "Animal — core",
        "out_of_scope": "Out of scope",
        "unknown": "Unknown",
    },
}

_WWF_FG_LABELS = {
    "FG1": "FG1 — Protéines",
    "FG2": "FG2 — Produits laitiers",
    "FG3": "FG3 — Matières grasses",
    "FG4": "FG4 — Fruits & légumes",
    "FG5": "FG5 — Céréales",
    "FG6": "FG6 — Féculents",
    "FG7": "FG7 — Snacks",
    "out_of_scope": "Hors périmètre",
    "unknown": "Inconnu",
}

# WWF Step-1 composite buckets — localised so the export reads like the app
# UI ("Composite · Végane") rather than a raw enum value ("vegan").
_WWF_BUCKET_LABELS = {
    "fr": {
        "meat_based": "À base de viande",
        "seafood_based": "À base de poisson",
        "vegetarian": "Végétarien",
        "vegan": "Végane",
    },
    "en": {
        "meat_based": "Meat-based",
        "seafood_based": "Seafood-based",
        "vegetarian": "Vegetarian",
        "vegan": "Vegan",
    },
}

_PT_ORDER = [
    "plant_based_core",
    "plant_based_non_core",
    "composite_products",
    "animal_core",
    "out_of_scope",
    "unknown",
]
_WWF_ORDER = ["FG1", "FG2", "FG3", "FG4", "FG5", "FG6", "FG7", "out_of_scope", "unknown"]

_HEADER_FILL = PatternFill("solid", fgColor="1F6B4C")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1F6B4C")


def _style_header(ws: Worksheet, ncols: int, row: int = 1) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _pct(v: float | None) -> str:
    return "" if v is None else f"{round(v * 100)} %"


def _kg(v: float | None) -> object:
    """Numeric kg rounded to 1 decimal (blank when unknown)."""
    return "" if v is None else round(v, 1)


def _plant_share(r: ExportRow) -> str:
    p, a = r.pt_plant_protein_kg, r.pt_animal_protein_kg
    if p is None or a is None:
        return ""
    total = p + a
    return "" if total <= 0 else f"{round(p / total * 100)} %"


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------


def _products_sheet(
    ws: Worksheet,
    rows: list[ExportRow],
    t: dict[str, str],
    pt_labels: dict[str, str],
    bucket_labels: dict[str, str],
    *,
    pt_enabled: bool,
    wwf_enabled: bool,
) -> None:
    # Per-product protein columns are only shown when a PT calculation has
    # actually produced amounts (otherwise they'd be a column of blanks).
    has_protein = pt_enabled and any(
        r.pt_total_protein_kg is not None for r in rows
    )
    headers = [t["h_id"], t["h_name"], t["h_retailer"]]
    if pt_enabled:
        headers += [t["h_pt_group"], t["h_pt_source"], t["h_pt_conf"]]
        if has_protein:
            headers += [
                t["h_pt_plant_kg"],
                t["h_pt_animal_kg"],
                t["h_pt_plant_share"],
            ]
    if wwf_enabled:
        headers += [
            t["h_wwf_group"],
            t["h_wwf_sub"],
            t["h_wwf_comp"],
            t["h_wwf_source"],
            t["h_wwf_conf"],
        ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in rows:
        line: list[object] = [
            r.external_product_id,
            r.product_name,
            r.retailer_category or "",
        ]
        if pt_enabled:
            line += [
                pt_labels.get(r.pt_group or "", r.pt_group or ""),
                r.pt_source or "",
                _pct(r.pt_confidence),
            ]
            if has_protein:
                line += [
                    _kg(r.pt_plant_protein_kg),
                    _kg(r.pt_animal_protein_kg),
                    _plant_share(r),
                ]
        if wwf_enabled:
            # Composite products (a non-null Step-1 bucket) are shown as
            # "Composite" in the food-group column — never only the
            # schema-filler food group — and their bucket carries the detail.
            is_composite = r.wwf_composite_bucket is not None
            wwf_group_label = (
                t["composite"]
                if is_composite
                else _WWF_FG_LABELS.get(
                    r.wwf_food_group or "", r.wwf_food_group or ""
                )
            )
            bucket = r.wwf_composite_bucket
            line += [
                wwf_group_label,
                "" if is_composite else (r.wwf_subgroup or ""),
                bucket_labels.get(bucket or "", bucket or "") if bucket else "",
                r.wwf_source or "",
                _pct(r.wwf_confidence),
            ]
        ws.append(line)
    widths = [16, 34, 22]
    if pt_enabled:
        widths += [18, 14, 12]
        if has_protein:
            widths += [20, 20, 16]
    if wwf_enabled:
        widths += [22, 22, 18, 14, 14]
    _autosize(ws, widths)
    ws.freeze_panes = "A2"


def _pt_analysis_sheet(
    ws: Worksheet, rows: list[ExportRow], t: dict[str, str], pt_labels: dict[str, str]
) -> None:
    counts = Counter(r.pt_group for r in rows if r.pt_group)
    present = [g for g in _PT_ORDER if counts.get(g)]
    ws.cell(row=1, column=1, value=t["a_pt_title"]).font = _TITLE_FONT
    ws.cell(row=3, column=1, value=t["a_category"])
    ws.cell(row=3, column=2, value=t["a_count"])
    _style_header(ws, 2, row=3)
    row = 4
    for g in present:
        ws.cell(row=row, column=1, value=pt_labels.get(g, g))
        ws.cell(row=row, column=2, value=counts[g])
        row += 1
    last = row - 1
    _autosize(ws, [26, 18])

    if not present:
        return

    # All data tables live in columns A–B near the top; every chart floats in
    # column D and below, each in its OWN row band so they never overlap (a
    # ~20-row gap clears an 8 cm chart at default row height).
    chart = BarChart()
    chart.type = "col"
    chart.title = t["a_pt_chart"]
    chart.legend = None
    chart.add_data(
        Reference(ws, min_col=2, min_row=3, max_row=last), titles_from_data=True
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=last))
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, "D2")

    # Plant vs animal vs composite, by PRODUCT COUNT (pie).
    plant_n = counts.get("plant_based_core", 0) + counts.get(
        "plant_based_non_core", 0
    )
    animal_n = counts.get("animal_core", 0)
    composite_n = counts.get("composite_products", 0)
    other_n = counts.get("out_of_scope", 0) + counts.get("unknown", 0)
    count_start = last + 3
    ws.cell(row=count_start, column=1, value=t["a_pt_pie"]).font = _TITLE_FONT
    crow = count_start + 1
    cfirst = crow
    for label, val in (
        (t["plant"], plant_n),
        (t["animal"], animal_n),
        (t["composite"], composite_n),
        (t["other"], other_n),
    ):
        ws.cell(row=crow, column=1, value=label)
        ws.cell(row=crow, column=2, value=val)
        crow += 1
    count_pie = PieChart()
    count_pie.title = t["a_pt_pie"]
    count_pie.add_data(Reference(ws, min_col=2, min_row=cfirst, max_row=crow - 1))
    count_pie.set_categories(
        Reference(ws, min_col=1, min_row=cfirst, max_row=crow - 1)
    )
    count_pie.height = 8
    count_pie.width = 12
    ws.add_chart(count_pie, "D22")

    # Protein split in KG (pie) — only when a PT calculation produced amounts.
    plant_kg = sum((r.pt_plant_protein_kg or 0.0) for r in rows)
    animal_kg = sum((r.pt_animal_protein_kg or 0.0) for r in rows)
    if plant_kg > 0 or animal_kg > 0:
        prot_start = crow + 2
        ws.cell(
            row=prot_start, column=1, value=t["a_pt_protein_title"]
        ).font = _TITLE_FONT
        prow = prot_start + 1
        pfirst = prow
        for label, val in (
            (t["plant"], round(plant_kg, 1)),
            (t["animal"], round(animal_kg, 1)),
        ):
            ws.cell(row=prow, column=1, value=label)
            ws.cell(row=prow, column=2, value=val)
            prow += 1
        prot_pie = PieChart()
        prot_pie.title = t["a_pt_protein_pie"]
        prot_pie.add_data(Reference(ws, min_col=2, min_row=pfirst, max_row=prow - 1))
        prot_pie.set_categories(
            Reference(ws, min_col=1, min_row=pfirst, max_row=prow - 1)
        )
        prot_pie.height = 8
        prot_pie.width = 12
        ws.add_chart(prot_pie, "D42")


def _wwf_analysis_sheet(
    ws: Worksheet, rows: list[ExportRow], t: dict[str, str]
) -> None:
    # Composites are tallied as a single "Composite" category — NEVER under
    # their schema-filler food group. This mirrors the products sheet and the
    # PT analysis sheet (which carves out "composite_products"), and matches
    # the canonical WWF calculation, which excludes composites from the
    # per-food-group breakdown and routes them by Step-1 bucket instead.
    counts: Counter[str] = Counter()
    for r in rows:
        if r.wwf_composite_bucket is not None:
            counts["composite"] += 1
        elif r.wwf_food_group:
            counts[r.wwf_food_group] += 1
    order = [*_WWF_ORDER[:7], "composite", *_WWF_ORDER[7:]]
    present = [g for g in order if counts.get(g)]

    def _wwf_label(g: str) -> str:
        return t["composite"] if g == "composite" else _WWF_FG_LABELS.get(g, g)

    ws.cell(row=1, column=1, value=t["a_wwf_title"]).font = _TITLE_FONT
    ws.cell(row=3, column=1, value=t["a_category"])
    ws.cell(row=3, column=2, value=t["a_count"])
    _style_header(ws, 2, row=3)
    row = 4
    for g in present:
        ws.cell(row=row, column=1, value=_wwf_label(g))
        ws.cell(row=row, column=2, value=counts[g])
        row += 1
    last = row - 1
    _autosize(ws, [24, 18])
    if present:
        chart = BarChart()
        chart.type = "col"
        chart.title = t["a_wwf_chart"]
        chart.legend = None
        data = Reference(ws, min_col=2, min_row=3, max_row=last)
        cats = Reference(ws, min_col=1, min_row=4, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, "D3")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def build_categorized_workbook(
    *,
    project_name: str,
    rows: list[ExportRow],
    pt_enabled: bool,
    wwf_enabled: bool,
    lang: str = "fr",
) -> bytes:
    """Build the categorised export workbook and return the .xlsx bytes."""
    lang = lang if lang in _L else "fr"
    t = _L[lang]
    pt_labels = _PT_GROUP_LABELS[lang]
    bucket_labels = _WWF_BUCKET_LABELS[lang]

    wb = Workbook()
    ws_products = wb.active
    ws_products.title = t["sheet_products"][:31]
    _products_sheet(
        ws_products,
        rows,
        t,
        pt_labels,
        bucket_labels,
        pt_enabled=pt_enabled,
        wwf_enabled=wwf_enabled,
    )

    if pt_enabled:
        _pt_analysis_sheet(wb.create_sheet(t["sheet_pt"][:31]), rows, t, pt_labels)
    if wwf_enabled:
        _wwf_analysis_sheet(wb.create_sheet(t["sheet_wwf"][:31]), rows, t)

    wb.properties.title = f"{t['title']} — {project_name}"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

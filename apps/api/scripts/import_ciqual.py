#!/usr/bin/env python3
"""Import ANSES-CIQUAL 2025 food composition data into the ciqual_reference table.

Usage
-----
    uv run python apps/api/scripts/import_ciqual.py --path /path/to/ciqual.xlsx

The script reads the "food composition" sheet from the official CIQUAL Excel
file, normalises the data (comma-decimal → float, "< N" → below_detection),
and upserts rows into the ``ciqual_reference`` Postgres table.

The CIQUAL file is NOT committed to the repository. Download it from:
    https://ciqual.anses.fr/

Attribution (required for any output using CIQUAL values):
    Anses. 2025. Ciqual French food composition table. https://ciqual.anses.fr/

Environment variables
---------------------
SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY — required for DB writes.
CIQUAL_DRY_RUN=1 — print rows instead of writing to DB.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

_BELOW_RE = re.compile(r"^<\s*([\d,]+)$")
_VERSION = "2025"

# Column indices in the CIQUAL Excel "food composition" sheet (0-based).
_COL_GRP_CODE = 0
_COL_SSGRP_CODE = 1
_COL_SSSSGRP_CODE = 2
_COL_GRP_NOM = 3
_COL_SSGRP_NOM = 4
_COL_SSSSGRP_NOM = 5
_COL_CODE = 6
_COL_NOM = 7
_COL_PROTEIN = 14  # "Protein (g 100g)"


def _parse_numeric(raw: object) -> tuple[Decimal | None, bool]:
    """Return (value, is_below_detection).

    Handles:
      - None / empty → (None, False)
      - "-"           → (None, False)   not analysed
      - "4,41"        → (Decimal("4.41"), False)
      - "< 0,2"       → (None, True)    below detection limit
      - 4.41          → (Decimal("4.41"), False)  already float from openpyxl
    """
    if raw is None:
        return None, False
    s = str(raw).strip()
    if s in ("", "-"):
        return None, False
    m = _BELOW_RE.match(s)
    if m:
        return None, True
    # Normalize comma-decimal
    s = s.replace(",", ".")
    try:
        return Decimal(s), False
    except InvalidOperation:
        return None, False


def _str_or_none(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s and s != "-" else None


def read_ciqual_excel(path: Path, *, verbose: bool = False) -> list[dict]:
    """Parse the CIQUAL Excel file and return a list of row dicts."""
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        print("ERROR: openpyxl is required. Run: uv add openpyxl --dev", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "food composition" not in wb.sheetnames:
        print(f"ERROR: sheet 'food composition' not found in {path}", file=sys.stderr)
        sys.exit(1)
    ws = wb["food composition"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Skip header row
    data_rows = rows[1:]
    if verbose:
        print(f"  {len(data_rows)} data rows found in {path.name}")

    entries: list[dict] = []
    skipped = 0
    for row in data_rows:
        if not row or row[_COL_CODE] is None:
            skipped += 1
            continue

        protein, below = _parse_numeric(row[_COL_PROTEIN])

        entries.append(
            {
                "id": str(uuid4()),
                "source": "ciqual",
                "source_version": _VERSION,
                "source_food_code": str(row[_COL_CODE]).strip(),
                "food_name_en": str(row[_COL_NOM] or "").strip(),
                "food_group": _str_or_none(row[_COL_GRP_NOM]) or "unknown",
                "food_subgroup": _str_or_none(row[_COL_SSGRP_NOM]),
                "food_subsubgroup": _str_or_none(row[_COL_SSSSGRP_NOM]),
                "protein_g_per_100g": float(protein) if protein is not None else None,
                "is_below_detection": below,
            }
        )

    if verbose and skipped:
        print(f"  {skipped} blank rows skipped")

    return entries


def upsert_to_db(entries: list[dict], *, dry_run: bool = False) -> None:
    """Upsert CIQUAL entries into the ciqual_reference Postgres table."""
    if dry_run:
        print(f"[DRY RUN] Would upsert {len(entries)} rows into ciqual_reference")
        for e in entries[:3]:
            print(f"  {e['source_food_code']:>8}  {e['food_name_en'][:50]:<50}  "
                  f"protein={e['protein_g_per_100g']}")
        if len(entries) > 3:
            print(f"  … and {len(entries) - 3} more")
        return

    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        print(
            "ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set for DB writes.\n"
            "       Set CIQUAL_DRY_RUN=1 to skip DB and print rows instead.",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        from supabase import create_client  # type: ignore[import-untyped]
    except ImportError:
        print("ERROR: supabase-py is required.", file=sys.stderr)
        sys.exit(1)

    client = create_client(url, key)

    batch_size = 500
    total_upserted = 0
    for i in range(0, len(entries), batch_size):
        batch = entries[i : i + batch_size]
        client.table("ciqual_reference").upsert(
            batch,
            on_conflict="source_version,source_food_code",
        ).execute()
        total_upserted += len(batch)
        print(f"  Upserted {total_upserted}/{len(entries)} rows…")

    print(f"Done. {total_upserted} rows upserted into ciqual_reference.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import ANSES-CIQUAL 2025 data into ciqual_reference table.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--path",
        type=Path,
        required=True,
        help="Path to the CIQUAL Excel file (.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=os.environ.get("CIQUAL_DRY_RUN") == "1",
        help="Print rows without writing to DB",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    if not args.path.is_file():
        print(f"ERROR: file not found: {args.path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {args.path}…")
    entries = read_ciqual_excel(args.path, verbose=args.verbose)
    print(f"Parsed {len(entries)} entries (version={_VERSION})")

    upsert_to_db(entries, dry_run=args.dry_run)


if __name__ == "__main__":
    main()

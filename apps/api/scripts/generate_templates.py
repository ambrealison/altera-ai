#!/usr/bin/env python
"""Phase Product-UX-B — generate import templates (CSV + XLSX).

Writes six static assets into ``apps/web/public/templates/`` (served by
Next.js): Protein Tracker, WWF, and combined, each as ``.csv`` and a
styled ``.xlsx``. Headers are chosen to match the parser's auto-mapping
synonyms exactly; the script asserts each set maps with no
missing-required field and no duplicate canonical before writing, so a
downloaded template always imports cleanly.

Run:
    .venv/bin/python scripts/generate_templates.py
"""

from __future__ import annotations

import collections
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# __file__ = apps/api/scripts/generate_templates.py → parents[2] = apps/
_APPS = Path(__file__).resolve().parents[2]
_OUT = _APPS / "web" / "public" / "templates"

# --- Header sets (parser-synonym-compatible) + example rows ---------------
PT = {
    "headers": [
        "product_id", "product_name", "raw_product_category",
        "ingredient_declaration_simulated", "pack_weight_g", "unit_purchased",
        "protein_total_g_per_100g", "protein_plant_g_per_100g",
        "protein_animal_g_per_100g", "protein_split_known", "label_claims_notes",
    ],
    "rows": [
        ["SKU-0001", "Lentilles vertes 500g", "Épicerie / Légumes secs",
         "Lentilles vertes", "500", "1200", "9.0", "9.0", "0.0", "true", "Bio"],
        ["SKU-0002", "Filet de poulet 400g", "Boucherie / Volaille",
         "Filet de poulet", "400", "850", "23.0", "0.0", "23.0", "true", ""],
    ],
}
WWF = {
    "headers": [
        "product_id", "product_name", "brand_type", "retail_channel",
        "raw_product_category", "ingredient_declaration_simulated",
        "pack_weight_g", "units_sold", "sales_weight_kg",
        "drained_weight_g_if_applicable", "label_claims_notes",
    ],
    "rows": [
        ["SKU-0001", "Lentilles vertes 500g", "Own brand", "grocery_ambient",
         "Épicerie / Légumes secs", "Lentilles vertes", "500", "1200", "600",
         "", "Bio"],
        ["SKU-0002", "Filet de saumon 200g", "National brand", "fresh",
         "Marée / Poisson frais", "Saumon atlantique", "200", "640", "128",
         "", ""],
    ],
}
COMBINED = {
    "headers": [
        "product_id", "product_name", "brand_type", "retail_channel",
        "raw_product_category", "ingredient_declaration_simulated",
        "pack_weight_g", "units_sold", "unit_purchased", "sales_weight_kg",
        "drained_weight_g_if_applicable", "protein_total_g_per_100g",
        "protein_plant_g_per_100g", "protein_animal_g_per_100g",
        "protein_split_known", "main_ingredient_origin_hint", "label_claims_notes",
    ],
    "rows": [
        ["SKU-0001", "Lentilles vertes 500g", "Own brand", "grocery_ambient",
         "Épicerie / Légumes secs", "Lentilles vertes", "500", "1200", "1200",
         "600", "", "9.0", "9.0", "0.0", "true", "FR", "Bio"],
        ["SKU-0002", "Yaourt nature 4x125g", "National brand", "fresh",
         "Crèmerie / Yaourts", "Lait, ferments lactiques", "500", "980", "980",
         "490", "", "4.0", "0.0", "4.0", "true", "FR", ""],
    ],
}

_HEADER_FILL = PatternFill("solid", fgColor="2C9D6B")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_csv(path: Path, spec: dict) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(spec["headers"])
        w.writerows(spec["rows"])


def _write_xlsx(path: Path, spec: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(spec["headers"])
    for row in spec["rows"]:
        ws.append(row)
    # Styled, frozen header + sensible column widths.
    for col_idx, header in enumerate(spec["headers"], start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        width = max(12, min(34, len(header) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def _verify(spec: dict, methodologies: list[str], label: str) -> None:
    from altera_api.ingestion.mapping import infer_mapping

    r = infer_mapping(spec["headers"], methodologies=methodologies)
    canon = [e.canonical_field for e in r.entries if e.canonical_field]
    dups = [k for k, v in collections.Counter(canon).items() if v > 1]
    problems = []
    if "protein_tracker" in methodologies and r.missing_required_pt:
        problems.append(f"missing PT {r.missing_required_pt}")
    if "wwf" in methodologies and r.missing_required_wwf:
        problems.append(f"missing WWF {r.missing_required_wwf}")
    if dups:
        problems.append(f"duplicate canonicals {dups}")
    if problems:
        raise SystemExit(f"{label} template invalid: {'; '.join(problems)}")


def main() -> int:
    sys.path.insert(0, str(_APPS / "api"))
    _OUT.mkdir(parents=True, exist_ok=True)
    targets = [
        ("protein_tracker", PT, ["protein_tracker"]),
        ("wwf", WWF, ["wwf"]),
        ("combined", COMBINED, ["protein_tracker", "wwf"]),
    ]
    for name, spec, methos in targets:
        _verify(spec, methos, name)
        _write_csv(_OUT / f"altera_template_{name}.csv", spec)
        _write_xlsx(_OUT / f"altera_template_{name}.xlsx", spec)
        print(f"  ✓ {name}: csv + xlsx ({len(spec['headers'])} cols)")
    print(f"Wrote 6 templates to {_OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Import RIVM NEVO-Online 2025 v9.0 food composition data into the nevo_reference table.

Usage
-----
    uv run python apps/api/scripts/import_nevo.py --path /path/to/NEVO2025_v9.0.csv

The script reads the pipe-delimited NEVO CSV export, normalises the data
(comma-decimal → Decimal, blank → None), and upserts rows into the
``nevo_reference`` Postgres table.

NEVO is preferred over CIQUAL for Protein Tracker because it publishes
plant protein (PROTPL) and animal protein (PROTAN) per 100 g.

The NEVO file is NOT committed to the repository. Download from:
    https://nevo-online.rivm.nl/

Attribution (required for any output using NEVO values):
    RIVM. 2025. NEVO-Online 2025 v9.0. https://nevo-online.rivm.nl/

Environment variables
---------------------
SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY — required for DB writes.
NEVO_DRY_RUN=1 — print rows instead of writing to DB.
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

_VERSION = "2025_v9.0"
_DELIMITER = "|"

# Column names from the NEVO CSV header (Dutch/English bilingual labels).
_COL_FOOD_GROUP_EN = "Food group"
_COL_CODE = "NEVO-code"
_COL_NAME_NL = "Voedingsmiddelnaam/Dutch food name"
_COL_NAME_EN = "Engelse naam/Food name"
_COL_QUANTITY = "Hoeveelheid/Quantity"
_COL_PROT = "PROT (g)"
_COL_PROTPL = "PROTPL (g)"
_COL_PROTAN = "PROTAN (g)"


def _parse_decimal(raw: object) -> Decimal | None:
    """Parse NEVO numeric cell (comma decimal). Blank/empty → None."""
    if raw is None:
        return None
    s = str(raw).strip().strip('"')
    if s == "":
        return None
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _str_clean(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip().strip('"')


def _row_to_entry(
    row: dict, *, rejected_counter: list[int], verbose: bool = False
) -> dict | None:
    """Build one nevo_reference upsert dict from a parsed row.

    Returns None for rows whose nevo_code is blank (NEVO header/spacer
    rows in Excel exports). Increments ``rejected_counter[0]`` when a
    protein value is negative — the entry is still emitted but all
    three protein fields are nulled, mirroring the CSV path.
    """
    code = _str_clean(row.get(_COL_CODE))
    if not code:
        return None

    prot = _parse_decimal(row.get(_COL_PROT))
    protpl = _parse_decimal(row.get(_COL_PROTPL))
    protan = _parse_decimal(row.get(_COL_PROTAN))

    # Non-negative validation. Bad rows are kept with NULL protein so the
    # check-constraint in nevo_reference doesn't reject the upsert; the
    # negative value itself is dropped on the floor.
    for label, value in (("PROT", prot), ("PROTPL", protpl), ("PROTAN", protan)):
        if value is not None and value < 0:
            rejected_counter[0] += 1
            if verbose:
                print(
                    f"  rejecting NEVO-code={code}: {label}={value} is negative",
                    file=sys.stderr,
                )
            prot = None
            protpl = None
            protan = None
            break

    return {
        "id": str(uuid4()),
        "source": "nevo",
        "source_version": _VERSION,
        "nevo_code": code,
        "food_name_nl": _str_clean(row.get(_COL_NAME_NL)),
        "food_name_en": _str_clean(row.get(_COL_NAME_EN)),
        "food_group": _str_clean(row.get(_COL_FOOD_GROUP_EN)) or "unknown",
        "quantity_basis": _str_clean(row.get(_COL_QUANTITY)) or "per 100g",
        "protein_g_per_100g": float(prot) if prot is not None else None,
        "plant_protein_g_per_100g": float(protpl) if protpl is not None else None,
        "animal_protein_g_per_100g": float(protan) if protan is not None else None,
    }


def read_nevo_csv(path: Path, *, verbose: bool = False) -> list[dict]:
    """Parse a pipe-delimited NEVO CSV export and return row dicts."""
    text = path.read_text(encoding="utf-8")
    reader = csv.DictReader(io.StringIO(text), delimiter=_DELIMITER)
    required = (
        _COL_CODE,
        _COL_NAME_NL,
        _COL_NAME_EN,
        _COL_FOOD_GROUP_EN,
        _COL_PROT,
    )
    missing = [c for c in required if c not in (reader.fieldnames or [])]
    if missing:
        raise ValueError(
            f"NEVO CSV missing required columns: {missing}; got {reader.fieldnames!r}"
        )

    entries: list[dict] = []
    skipped = 0
    rejected_counter = [0]
    for row in reader:
        entry = _row_to_entry(row, rejected_counter=rejected_counter, verbose=verbose)
        if entry is None:
            skipped += 1
            continue
        entries.append(entry)

    if verbose:
        print(f"  {len(entries)} entries parsed, {skipped} blank rows skipped")
        if rejected_counter[0]:
            print(f"  {rejected_counter[0]} entries had negative protein values (nulled)")

    return entries


def read_nevo_excel(path: Path, *, verbose: bool = False) -> list[dict]:
    """Parse a NEVO 2025 .xlsx workbook (sheet ``NEVO2025``) and return rows.

    Excel cells may arrive as floats, ints, or strings depending on how
    NEVO published them. ``_parse_decimal`` already handles the
    comma-decimal string case; numeric cells are passed through ``str()``
    by ``_str_clean`` and then re-parsed, which round-trips cleanly.
    """
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        print(
            "ERROR: openpyxl is required for .xlsx. Run: uv add openpyxl --dev",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_name = "NEVO2025" if "NEVO2025" in wb.sheetnames else wb.sheetnames[0]
    if verbose:
        print(f"  reading sheet {sheet_name!r}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        raise ValueError(f"NEVO workbook sheet {sheet_name!r} is empty")

    # Normalise header cells (some exports have quotes/whitespace).
    header_labels = [str(h).strip().strip('"') if h is not None else "" for h in header]
    required = (
        _COL_CODE,
        _COL_NAME_NL,
        _COL_NAME_EN,
        _COL_FOOD_GROUP_EN,
        _COL_PROT,
    )
    missing = [c for c in required if c not in header_labels]
    if missing:
        wb.close()
        raise ValueError(
            f"NEVO workbook missing required columns: {missing}; got {header_labels!r}"
        )

    entries: list[dict] = []
    skipped = 0
    rejected_counter = [0]
    for row in rows_iter:
        if row is None:
            skipped += 1
            continue
        row_dict = dict(zip(header_labels, row, strict=False))
        entry = _row_to_entry(row_dict, rejected_counter=rejected_counter, verbose=verbose)
        if entry is None:
            skipped += 1
            continue
        entries.append(entry)
    wb.close()

    if verbose:
        print(f"  {len(entries)} entries parsed, {skipped} blank rows skipped")
        if rejected_counter[0]:
            print(f"  {rejected_counter[0]} entries had negative protein values (nulled)")

    return entries


def read_nevo(path: Path, *, verbose: bool = False) -> list[dict]:
    """Dispatch by file extension."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return read_nevo_excel(path, verbose=verbose)
    return read_nevo_csv(path, verbose=verbose)


def upsert_to_db(entries: list[dict], *, dry_run: bool = False) -> None:
    """Upsert NEVO entries into the nevo_reference Postgres table."""
    if dry_run:
        print(f"[DRY RUN] Would upsert {len(entries)} rows into nevo_reference")
        for e in entries[:3]:
            split = (
                f"plant={e['plant_protein_g_per_100g']}, animal={e['animal_protein_g_per_100g']}"
                if e["plant_protein_g_per_100g"] is not None
                else "no split"
            )
            print(
                f"  {e['nevo_code']:>6}  {e['food_name_en'][:50]:<50}  "
                f"prot={e['protein_g_per_100g']}  ({split})"
            )
        if len(entries) > 3:
            print(f"  … and {len(entries) - 3} more")
        return

    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        print(
            "ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set for DB writes.\n"
            "       Set NEVO_DRY_RUN=1 to skip DB and print rows instead.",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        from supabase import create_client  # type: ignore[import-untyped]
    except ImportError:
        print("ERROR: supabase-py is required.", file=sys.stderr)
        sys.exit(1)

    client = create_client(url, key)

    batch_size = 500
    total_upserted = 0
    for i in range(0, len(entries), batch_size):
        batch = entries[i : i + batch_size]
        client.table("nevo_reference").upsert(
            batch,
            on_conflict="source_version,nevo_code",
        ).execute()
        total_upserted += len(batch)
        print(f"  Upserted {total_upserted}/{len(entries)} rows…")

    print(f"Done. {total_upserted} rows upserted into nevo_reference.")


#: Phase 34N — NEVO 2025 v9.0 is expected to ship ~2,328 entries.
#: Below this floor the importer fails loudly so the operator notices
#: a truncated upload rather than silently shipping 1000 rows that
#: would later cap matching coverage.
_EXPECTED_MIN_ROWS = 2000


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import RIVM NEVO-Online 2025 v9.0 data into nevo_reference table.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--path",
        type=Path,
        required=True,
        help=(
            "Path to the NEVO source file. Pipe-delimited .csv or .xlsx "
            "workbook (sheet 'NEVO2025'). Dispatch is by file extension."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=os.environ.get("NEVO_DRY_RUN") == "1",
        help="Print rows without writing to DB",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help=(
            f"Phase 34N — truncate the imported entries to N rows. Only "
            f"use this for tests / smoke imports; production imports "
            f"MUST process the full dataset. Without --limit the "
            f"importer fails if fewer than {_EXPECTED_MIN_ROWS} rows were parsed."
        ),
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    if not args.path.is_file():
        print(f"ERROR: file not found: {args.path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {args.path}…")
    entries = read_nevo(args.path, verbose=args.verbose)
    print(f"Parsed {len(entries)} entries (version={_VERSION})")

    if args.limit is not None and args.limit > 0:
        entries = entries[: args.limit]
        print(f"--limit applied — truncated to {len(entries)} entries")
    elif len(entries) < _EXPECTED_MIN_ROWS:
        print(
            f"ERROR: NEVO 2025 import returned only {len(entries)} rows "
            f"(expected >= {_EXPECTED_MIN_ROWS}). This usually means the "
            f"source file is truncated or the wrong file was downloaded. "
            f"Re-download from https://nevo-online.rivm.nl/ or pass "
            f"--limit N if a smaller dataset is intentional.",
            file=sys.stderr,
        )
        sys.exit(2)

    upsert_to_db(entries, dry_run=args.dry_run)


if __name__ == "__main__":
    main()

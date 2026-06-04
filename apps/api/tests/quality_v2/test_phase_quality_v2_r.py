"""Phase Quality-V2-R — internal review workflow package for NEVO V2 dry-run.

Annotates each dry-run proposal with a suggested_action, a review_priority
(P0..P3) and a review_bucket, writes a consolidated review workbook (xlsx, or a
CSV fallback), and surfaces the counts in the JSON summary. Pure annotation +
artifacts: the matcher and the safety actions are unchanged, and there are no
DB writes / no routes / no Supabase.
"""

from __future__ import annotations

import builtins
import csv
import json
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest

from altera_api.classification_v2 import nevo_v2_enrich as cli
from altera_api.classification_v2.nevo_matcher import NevoMatcherVersion
from altera_api.classification_v2.nevo_review_workflow import (
    MANUAL_DECISION_VALUES,
    REVIEW_BUCKETS,
    REVIEW_PRIORITIES,
    SUGGESTED_ACTIONS,
    classify_product_policy,
    review_priority,
    suggested_action,
)


def _row(name, action, *, outcome="match", conf=0.95, reason="",
         top5="Cand A | Cand B"):
    return {
        "product_name": name, "nutrition_safety_action": action,
        "matcher_outcome": outcome, "matcher_confidence": conf,
        "nutrition_safety_reason": reason, "top_5_candidates": top5,
    }


# ---------------------------------------------------------------------------
# Part B — suggested_action mapping.
# ---------------------------------------------------------------------------
class TestSuggestedAction:
    @pytest.mark.parametrize(
        "row,expected",
        [
            (_row("Chocolat Noir", "would_enrich"), "approve_auto_candidate"),
            (_row("Pâtes Fusilli", "skip_state_mismatch"),
             "review_state_mismatch"),
            (_row("Compote Pomme", "skip_proxy_too_broad",
                  reason="reference is a processing proxy (syrup)"),
             "review_proxy_too_broad"),
            (_row("Chips Vinaigre", "route_to_review",
                  reason="generic snack proxy: ..."), "review_generic_proxy"),
            (_row("Huile Colza", "route_to_review",
                  reason="branded blend/margarine-like spread"),
             "review_generic_proxy"),
            (_row("Truc Food", "route_to_review",
                  reason="matcher result is review-level / low-confidence"),
             "needs_manual_nevo_search"),
            (_row("Specialite", "skip_no_match", outcome="no_match",
                  top5="Cand A | Cand B"), "review_no_match"),
            (_row("Specialite", "skip_no_match", outcome="no_match", top5=""),
             "needs_manual_nevo_search"),
        ],
    )
    def test_mapping(self, row, expected) -> None:
        assert suggested_action(row) == expected

    def test_non_food_and_pet(self) -> None:
        assert suggested_action(
            _row("Liquide Vaisselle Citron", "skip_no_match",
                 outcome="no_match", top5="")
        ) == "reject_non_food"
        assert suggested_action(
            _row("Litière Chat Agglomérante", "skip_no_match",
                 outcome="no_match", top5="")
        ) == "reject_policy_excluded"

    def test_every_action_is_declared(self) -> None:
        # Sanity: the mapping never invents an undeclared value.
        rows = [
            _row("Chocolat Noir", "would_enrich"),
            _row("Pâtes", "skip_state_mismatch"),
            _row("Compote", "skip_proxy_too_broad"),
            _row("Chips Sel", "route_to_review", reason="generic snack proxy"),
            _row("X Food", "skip_no_match", outcome="no_match"),
            _row("Shampooing", "skip_no_match", outcome="no_match", top5=""),
        ]
        assert {suggested_action(r) for r in rows} <= set(SUGGESTED_ACTIONS)


# ---------------------------------------------------------------------------
# Part B — review_priority mapping.
# ---------------------------------------------------------------------------
class TestReviewPriority:
    def test_would_enrich_high_vs_low_confidence(self) -> None:
        assert review_priority(_row("Choc", "would_enrich", conf=0.99)) == "P2"
        assert review_priority(_row("Choc", "would_enrich", conf=0.92)) == "P1"

    def test_review_buckets_are_p1(self) -> None:
        for action, reason in (
            ("skip_state_mismatch", ""),
            ("skip_proxy_too_broad", "wrong vinegar type"),
            ("route_to_review", "generic snack proxy"),
        ):
            assert review_priority(_row("Food", action, reason=reason)) == "P1"

    def test_non_food_no_match_is_p3(self) -> None:
        assert review_priority(
            _row("Shampooing Anti-Pelliculaire", "skip_no_match",
                 outcome="no_match", top5="")
        ) == "P3"

    def test_non_food_matched_is_p0(self) -> None:
        # A non-food the matcher ACCEPTED is the most dangerous → never auto.
        assert review_priority(
            _row("Liquide Vaisselle Citron", "would_enrich", outcome="match",
                 conf=0.98)
        ) == "P0"

    def test_food_no_match_manual_search_priority(self) -> None:
        assert review_priority(
            _row("Specialite", "skip_no_match", outcome="no_match",
                 top5="Cand A | Cand B")
        ) == "P1"
        assert review_priority(
            _row("Specialite", "skip_no_match", outcome="no_match", top5="")
        ) == "P2"

    def test_priorities_are_declared(self) -> None:
        assert set(REVIEW_PRIORITIES) == {"P0", "P1", "P2", "P3"}


class TestPolicyClassification:
    def test_food_concept_wins(self) -> None:
        assert classify_product_policy("Chocolat Noir 70%") == "food"
        # Unknown but not flagged → treated as food (route to manual search).
        assert classify_product_policy("Specialite Inconnue XYZ") == "food"

    def test_pet_and_non_food(self) -> None:
        assert classify_product_policy("Pâtée Chien Boeuf Légumes") == "pet"
        assert classify_product_policy("Litière Chat") == "pet"
        assert classify_product_policy("Liquide Vaisselle Citron") == "non_food"
        assert classify_product_policy("Shampooing Anti-Pelliculaire") == (
            "non_food"
        )


# ---------------------------------------------------------------------------
# Proposal building + review package (Part A/C).
# ---------------------------------------------------------------------------
def _decision(*, matched, code, food, conf, review):
    return SimpleNamespace(
        matched=matched, nevo_code=code, food_name_en=food, confidence=conf,
        match_type="concept", review_required=review,
        top_candidates=[SimpleNamespace(candidate_name=food or "x",
                                        rejection_reason="")],
    )


class _FakeMatcher:
    def __init__(self, by_name):
        self._by_name = by_name

    def decide(self, query, top_k):  # noqa: ARG002
        return self._by_name[query["product_name"]]


def _p(name):
    return SimpleNamespace(id=uuid4(), product_name=name,
                           retailer_category=None, ingredients_text=None,
                           labels=())


def _entry(code, protein, name):
    return SimpleNamespace(nevo_code=code, protein_g_per_100g=protein,
                           food_name_en=name)


def _mixed_rows():
    products = [
        _p("Chocolat Noir 70%"),          # approve_auto_candidate
        _p("Pâtes Fusilli Blé Complet"),  # review_state_mismatch
        _p("Liquide Vaisselle Citron"),   # reject_non_food (no_match)
        _p("Specialite Inconnue XYZ"),    # needs_manual / review_no_match
    ]
    decisions = {
        "Chocolat Noir 70%": _decision(matched=True, code="C",
                                       food="Chocolate dark", conf=0.99,
                                       review=False),
        "Pâtes Fusilli Blé Complet": _decision(matched=True, code="P",
                                               food="Pasta wholemeal boiled",
                                               conf=0.95, review=False),
        "Liquide Vaisselle Citron": _decision(matched=False, code="", food="",
                                              conf=0.0, review=True),
        "Specialite Inconnue XYZ": _decision(matched=False, code="", food="",
                                            conf=0.0, review=True),
    }
    nevo_by_code = {
        "C": _entry("C", 7.0, "Chocolate dark"),
        "P": _entry("P", 5.0, "Pasta wholemeal boiled"),
    }
    return cli.build_proposals(
        products, version=NevoMatcherVersion.V2_EMBEDDINGS,
        matcher=_FakeMatcher(decisions), nevo=None, nevo_by_code=nevo_by_code,
        provider_name="fake", model="m", top_k=5,
    )


class TestReviewerColumns:
    def test_filtered_csv_has_workflow_and_reviewer_columns(self, tmp_path) -> None:
        rows = _mixed_rows()
        artifacts = cli.write_filtered_review_csvs(tmp_path, "demo", rows)
        path = Path(artifacts["would_enrich"]["path"])
        with path.open(encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            header = reader.fieldnames or []
            data = list(reader)
        for col in ("manual_decision", "reviewer_notes", "approved_nevo_code",
                    "approved_nevo_name", "approved_protein_g_per_100g",
                    "review_priority", "suggested_action"):
            assert col in header
        # computed columns populated, reviewer columns blank.
        assert data
        for r in data:
            assert r["suggested_action"] == "approve_auto_candidate"
            assert r["review_priority"] in REVIEW_PRIORITIES
            assert r["manual_decision"] == ""
            assert r["approved_protein_g_per_100g"] == ""


class TestReviewWorkbook:
    def test_xlsx_workbook_created_with_tabs(self, tmp_path) -> None:
        from openpyxl import load_workbook

        rows = _mixed_rows()
        summary = cli.build_dry_run_summary(
            rows, project_id="demo", version="v2-embeddings", provider="fake",
            model="m", top_k=5, generated_at=None,
        )
        path = Path(cli.write_review_package(tmp_path, "demo", rows, summary))
        assert path.suffix == ".xlsx"
        assert path.exists()
        wb = load_workbook(path)
        assert {"Summary", "Auto_Ready", "Needs_Review", "State_Mismatch",
                "Proxy_Too_Broad", "No_Match", "Non_Food_Policy",
                "Instructions"} <= set(wb.sheetnames)
        # Auto_Ready holds the chocolate row; header carries review_bucket.
        auto = wb["Auto_Ready"]
        header = [c.value for c in auto[1]]
        assert "review_bucket" in header and "suggested_action" in header
        body = list(auto.iter_rows(min_row=2, values_only=True))
        assert len(body) == 1
        # Instructions tab documents the manual_decision vocabulary.
        instr = "\n".join(
            str(c[0].value) for c in wb["Instructions"].iter_rows()
        )
        for value in MANUAL_DECISION_VALUES:
            assert value in instr

    def test_csv_fallback_when_openpyxl_missing(self, tmp_path, monkeypatch) -> None:
        real_import = builtins.__import__

        def _no_openpyxl(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("simulated: openpyxl unavailable")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr(builtins, "__import__", _no_openpyxl)
        rows = _mixed_rows()
        summary = cli.build_dry_run_summary(
            rows, project_id="demo", version="v2-embeddings", provider="fake",
            model="m", top_k=5, generated_at=None,
        )
        path = Path(cli.write_review_package(tmp_path, "demo", rows, summary))
        assert path.suffix == ".csv"
        with path.open(encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            header = reader.fieldnames or []
            data = list(reader)
        assert "review_bucket" in header
        assert len(data) == len(rows)
        assert {r["review_bucket"] for r in data} <= set(REVIEW_BUCKETS)


class TestSummaryCounts:
    def test_summary_has_review_workflow_counts(self) -> None:
        rows = _mixed_rows()
        summary = cli.build_dry_run_summary(
            rows, project_id="demo", version="v2-embeddings", provider="fake",
            model="m", top_k=5, generated_at=None,
        )
        assert set(summary["review_bucket_counts"]) == set(REVIEW_BUCKETS)
        assert set(summary["suggested_action_counts"]) == set(SUGGESTED_ACTIONS)
        assert set(summary["review_priority_counts"]) == {"P0", "P1", "P2", "P3"}
        # totals are consistent with the row count.
        assert sum(summary["review_bucket_counts"].values()) == len(rows)
        assert sum(summary["review_priority_counts"].values()) == len(rows)
        assert summary["suggested_action_counts"]["reject_non_food"] == 1
        assert summary["instructions_summary"]


# ---------------------------------------------------------------------------
# End-to-end dry-run — package written, no DB writes.
# ---------------------------------------------------------------------------
class _ReadOnlyStore:
    _WRITES = frozenset({"add_enrichment_record", "upsert_pt_classification",
                         "add_product", "add_run", "add_job"})

    def __init__(self, products, nevo_entries=()):
        self._products = products
        self._nevo = list(nevo_entries)
        self.reads: list[str] = []

    def get_project(self, project_id):
        self.reads.append("get_project")
        return object()

    def list_products_for_project(self, project_id):
        self.reads.append("list_products_for_project")
        return self._products

    def list_nevo_entries(self):
        self.reads.append("list_nevo_entries")
        return self._nevo

    def __getattr__(self, name):
        if name in self._WRITES:
            raise AssertionError(f"read-only violation: {name}")
        raise AttributeError(name)


def _store_product(name):
    return SimpleNamespace(
        id=uuid4(), product_name=name, retailer_category=None,
        retailer_subcategory=None, ingredients_text=None, labels=(),
        pt_fields=object(),
    )


def _nevo_entry(code, protein, name):
    return SimpleNamespace(
        nevo_code=code, protein_g_per_100g=protein, food_name_en=name,
        food_name_nl="", food_group="", plant_protein_g_per_100g=None,
        animal_protein_g_per_100g=None,
    )


class TestDryRunEndToEnd:
    def test_review_package_written_and_no_db(
        self, tmp_path, monkeypatch, capsys
    ) -> None:
        monkeypatch.delenv("ALTERA_ENABLE_EMBEDDINGS", raising=False)
        store = _ReadOnlyStore(
            [_store_product("Tofu nature"),
             _store_product("Litière Chat Agglomérante")],
            nevo_entries=[_nevo_entry("NEVO-TOFU", 8.0, "Tofu")],
        )
        pid = str(uuid4())
        rc = cli.main(
            ["--project-id", pid, "--matcher-version", "v2-embeddings",
             "--evaluator-fake", "--reference-source", "fixture",
             "--cache-dir", "", "--output-dir", str(tmp_path)],
            store=store,
        )
        assert rc == 0
        summary = json.loads(
            (tmp_path / f"nevo_v2_enrich_proposals_{pid}.json").read_text()
        )
        pkg = Path(summary["review_package_path"])
        assert pkg.exists()
        assert pkg.suffix in (".xlsx", ".csv")
        assert "review_bucket_counts" in summary
        assert "review_priority_counts" in summary
        assert "suggested_action_counts" in summary
        assert summary["persisted_writes"] == 0
        out = capsys.readouterr().out
        assert "REVIEW PRIORITY" in out
        assert "Consolidated review package" in out
        assert set(store.reads) <= {
            "get_project", "list_products_for_project", "list_nevo_entries"
        }


# ---------------------------------------------------------------------------
# Safety — defaults unchanged; routes clean.
# ---------------------------------------------------------------------------
class TestSafety:
    def test_v1_default_and_embeddings_off(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from altera_api.classification_v2.nevo_matcher import (
            resolve_nevo_matcher_version,
        )
        from altera_api.quality_config import embeddings_enabled

        monkeypatch.delenv("ALTERA_NEVO_MATCHER_VERSION", raising=False)
        monkeypatch.delenv("ALTERA_ENABLE_EMBEDDINGS", raising=False)
        assert str(resolve_nevo_matcher_version()) == "v1"
        assert embeddings_enabled() is False

    def test_routes_do_not_import_v2(self) -> None:
        api_dir = Path(cli.__file__).resolve().parents[1] / "api"
        offenders = [
            p.name for p in api_dir.rglob("*.py")
            if "classification_v2" in p.read_text(encoding="utf-8")
            or "altera_api.embeddings" in p.read_text(encoding="utf-8")
            or "nevo_review_workflow" in p.read_text(encoding="utf-8")
        ]
        assert not offenders

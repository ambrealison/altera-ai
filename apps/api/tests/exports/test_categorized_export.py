"""Categorised retailer export — workbook builder + download route."""

from __future__ import annotations

from collections.abc import Iterator
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from altera_api.api.state import InMemoryStore
from altera_api.api.store_factory import get_store
from altera_api.domain.common import ClassificationSource, Methodology
from altera_api.domain.product import (
    NormalizedProduct,
    PTProductFields,
    RetailChannel,
    WWFProductFields,
)
from altera_api.domain.protein_tracker import (
    ProteinTrackerGroup,
    ProteinTrackerProductClassification,
)
from altera_api.domain.wwf import (
    WWFFG1Subgroup,
    WWFFoodGroup,
    WWFProductClassification,
)
from altera_api.exports.categorized_workbook import (
    ExportRow,
    build_categorized_workbook,
)
from altera_api.main import app


@pytest.fixture
def store() -> InMemoryStore:
    return InMemoryStore()


@pytest.fixture
def client(
    store: InMemoryStore, monkeypatch: pytest.MonkeyPatch
) -> Iterator[TestClient]:
    monkeypatch.setenv("ALTERA_DEV_AUTH_ENABLED", "true")
    monkeypatch.delenv("SUPABASE_JWT_SECRET", raising=False)
    app.dependency_overrides[get_store] = lambda: store
    try:
        with TestClient(app) as c:
            yield c
    finally:
        app.dependency_overrides.pop(get_store, None)


# ---------------------------------------------------------------------------
# Builder unit tests
# ---------------------------------------------------------------------------


def _rows() -> list[ExportRow]:
    # A genuine WWF composite row (a vegetarian prepared pizza, à la demo50's
    # PTWWF049) exercises the composite-display path: FG2/cheese here is only
    # the schema filler — the bucket carries the real detail. (Note: demo25's
    # vegan pizza is a plain FG5 grain, not a composite — see the demo golden
    # test; this fixture deliberately keeps a composite to guard that code.)
    return [
        ExportRow("PTWWF001", "Lentilles", "Légumineuses", "plant_based_core",
                  "deterministic", 1.0, "FG1", "legumes", None, "deterministic", 1.0),
        ExportRow("PTWWF007", "Steak bœuf", "Viande", "animal_core",
                  "deterministic", 1.0, "FG1", "red_meat", None, "deterministic", 1.0),
        ExportRow("PTWWF049", "Pizza fromage tomate", "Plat préparé",
                  "composite_products", "deterministic", 1.0, "FG2",
                  "cheese", "vegetarian", "deterministic", 1.0),
    ]


class TestWorkbookBuilder:
    def test_three_sheets_when_both_methodologies(self) -> None:
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True
        )
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == ["Produits", "Analyse Protein Tracker", "Analyse WWF"]

    def test_pt_only_omits_wwf_sheet(self) -> None:
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=False
        )
        wb = load_workbook(BytesIO(data))
        assert "Analyse WWF" not in wb.sheetnames
        assert "Analyse Protein Tracker" in wb.sheetnames

    def test_english_sheet_names(self) -> None:
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True, lang="en"
        )
        wb = load_workbook(BytesIO(data))
        assert wb.sheetnames == ["Products", "Protein Tracker analysis", "WWF analysis"]

    def test_products_sheet_has_a_row_per_product(self) -> None:
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True
        )
        ws = load_workbook(BytesIO(data))["Produits"]
        # header + 3 products
        assert ws.max_row == 4
        names = [ws.cell(row=r, column=2).value for r in range(2, 5)]
        assert names == ["Lentilles", "Steak bœuf", "Pizza fromage tomate"]

    def test_charts_are_embedded(self) -> None:
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True
        )
        wb = load_workbook(BytesIO(data))
        assert len(wb["Analyse Protein Tracker"]._charts) >= 1
        assert len(wb["Analyse WWF"]._charts) >= 1

    def test_composite_shows_composite_not_food_group(self) -> None:
        # The Pizza row is a vegetarian composite (filler FG2). Its WWF group
        # cell must read "Composite", never the FG2 food-group label, and the
        # bucket column carries the LOCALISED bucket ("Végétarien"), never the
        # raw enum value or the filler subgroup.
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True
        )
        ws = load_workbook(BytesIO(data))["Produits"]
        by_id = {
            ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)
        }
        pizza = by_id["PTWWF049"]
        assert ws.cell(row=pizza, column=7).value == "Composite"  # WWF group
        assert "FG2" not in str(ws.cell(row=pizza, column=7).value)
        # subgroup hidden for composites (openpyxl reads "" back as None)
        assert ws.cell(row=pizza, column=8).value in (None, "")
        assert ws.cell(row=pizza, column=9).value == "Végétarien"  # localised bucket

    def test_wwf_analysis_counts_composite_separately_not_filler_fg(self) -> None:
        # Regression guard: the WWF analysis distribution must tally the
        # composite under a "Composite" category, NOT under its FG2 schema
        # filler. FG1 reflects only the two genuine FG1 products; the filler
        # FG2 must not appear at all.
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True
        )
        ws = load_workbook(BytesIO(data))["Analyse WWF"]
        dist: dict[str, int] = {}
        for r in range(4, ws.max_row + 1):
            cat = ws.cell(row=r, column=1).value
            cnt = ws.cell(row=r, column=2).value
            if cat is not None:
                dist[str(cat)] = cnt
        assert dist.get("Composite") == 1
        assert dist.get("FG1 — Protéines") == 2  # Lentilles + Steak, NOT the pizza
        # The composite's FG2 schema filler must NOT leak into the distribution.
        assert "FG2 — Produits laitiers" not in dist
        assert sum(dist.values()) == 3  # no double-counting


def _rows_with_protein() -> list[ExportRow]:
    return [
        ExportRow("PTWWF001", "Lentilles", "Lég", "plant_based_core",
                  "deterministic", 0.95, "FG1", "legumes", None,
                  "deterministic", 0.95, 200.0, 0.0, 200.0),
        ExportRow("PTWWF007", "Steak bœuf", "Viande", "animal_core",
                  "ai", 0.92, "FG1", "red_meat", None, "ai", 0.93,
                  0.0, 300.0, 300.0),
    ]


class TestProteinColumnsAndCharts:
    def test_products_sheet_has_protein_amounts_and_split(self) -> None:
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows_with_protein(),
            pt_enabled=True, wwf_enabled=True,
        )
        ws = load_workbook(BytesIO(data))["Produits"]
        hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Protéines végétales (kg)" in hdr
        assert "Protéines animales (kg)" in hdr
        assert "Part végétale (%)" in hdr
        # Lentilles: all-plant → 200 kg plant, 0 animal, 100 % plant share.
        plant_col = hdr.index("Protéines végétales (kg)") + 1
        share_col = hdr.index("Part végétale (%)") + 1
        assert ws.cell(row=2, column=plant_col).value == 200
        assert ws.cell(row=2, column=share_col).value == "100 %"

    def test_pt_sheet_has_three_separated_charts(self) -> None:
        # bar (groups) + count pie + protein-split pie, in distinct row bands.
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows_with_protein(),
            pt_enabled=True, wwf_enabled=True,
        )
        pt = load_workbook(BytesIO(data))["Analyse Protein Tracker"]
        assert len(pt._charts) == 3
        rows = sorted(c.anchor._from.row for c in pt._charts)
        # each chart band is clear of the previous (no overlap like before).
        assert all(rows[i + 1] - rows[i] >= 15 for i in range(len(rows) - 1))

    def test_no_protein_columns_or_chart_without_a_run(self) -> None:
        # _rows() carries no protein amounts (no calculation yet): the columns
        # and the protein chart are omitted, leaving bar + count pie only.
        data = build_categorized_workbook(
            project_name="Demo", rows=_rows(), pt_enabled=True, wwf_enabled=True
        )
        wb = load_workbook(BytesIO(data))
        hdr = [
            wb["Produits"].cell(row=1, column=c).value
            for c in range(1, wb["Produits"].max_column + 1)
        ]
        assert "Protéines végétales (kg)" not in hdr
        assert len(wb["Analyse Protein Tracker"]._charts) == 2


# ---------------------------------------------------------------------------
# Download route
# ---------------------------------------------------------------------------


def _seed_classified(store: InMemoryStore) -> str:
    project = store.create_project(
        name="Export demo",
        methodologies_enabled=frozenset(
            {Methodology.PROTEIN_TRACKER, Methodology.WWF}
        ),
        reporting_period_label="FY 2024",
        organisation_id=store.default_org_id,
        created_by=store.default_user_id,
    )
    upload_id = uuid4()
    now = datetime.now(UTC)
    pids: list = []
    for i in range(3):
        p = NormalizedProduct(
            id=uuid4(),
            project_id=project.id,
            upload_id=upload_id,
            organisation_id=store.default_org_id,
            row_number=i + 1,
            external_product_id=f"PTWWF{i + 1:03d}",
            product_name=f"Produit {i + 1}",
            brand=None,
            is_own_brand=False,
            retailer_category="Légumineuses",
            weight_per_item_kg=Decimal("0.15"),
            methodologies_enabled=frozenset(
                {Methodology.PROTEIN_TRACKER, Methodology.WWF}
            ),
            pt_fields=PTProductFields(items_purchased=Decimal("100")),
            wwf_fields=WWFProductFields(
                items_sold=Decimal("100"),
                retail_channel=RetailChannel.GROCERY_AMBIENT,
                is_own_brand=False,
            ),
            created_at=now,
        )
        store.add_product(p)
        pids.append(p.id)
        store.upsert_pt_classification(
            ProteinTrackerProductClassification(
                product_id=p.id,
                pt_group=ProteinTrackerGroup.PLANT_BASED_CORE,
                source=ClassificationSource.DETERMINISTIC,
                confidence=Decimal("1"),
                rule_id="test.rule",
                updated_at=now,
            )
        )
        store.upsert_wwf_classification(
            WWFProductClassification(
                product_id=p.id,
                wwf_food_group=WWFFoodGroup.FG1,
                wwf_is_composite=False,
                fg1_subgroup=WWFFG1Subgroup.LEGUMES,
                source=ClassificationSource.DETERMINISTIC,
                confidence=Decimal("1"),
                rule_id="test.rule",
                updated_at=now,
            )
        )
    return str(project.id)


class TestExportRoute:
    def test_download_returns_xlsx(
        self, client: TestClient, store: InMemoryStore
    ) -> None:
        project_id = _seed_classified(store)
        r = client.get(f"/api/v1/projects/{project_id}/export/categorized.xlsx")
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]
        assert "attachment" in r.headers.get("content-disposition", "")
        wb = load_workbook(BytesIO(r.content))
        assert "Produits" in wb.sheetnames
        assert wb["Produits"].max_row == 4  # header + 3

    def test_english_lang_param(
        self, client: TestClient, store: InMemoryStore
    ) -> None:
        project_id = _seed_classified(store)
        r = client.get(
            f"/api/v1/projects/{project_id}/export/categorized.xlsx?lang=en"
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert "Products" in wb.sheetnames

    def test_protein_columns_appear_after_a_pt_run(
        self, client: TestClient, store: InMemoryStore
    ) -> None:
        # End-to-end: once a finished PT run exists, the route reads its
        # per-product protein from rows_payload and the Produits sheet gains
        # the protein columns (all-plant products → plant=protein, animal=0).
        from uuid import UUID

        from altera_api.api.state import RunRecord
        from altera_api.domain.protein_tracker import (
            ProteinTrackerCalculationRow,
        )

        project_id = _seed_classified(store)
        products = store.list_products_for_project(UUID(project_id))
        run_id = uuid4()
        now = datetime.now(UTC)
        rows = [
            ProteinTrackerCalculationRow(
                run_id=run_id,
                product_id=p.id,
                in_scope=True,
                pt_group=ProteinTrackerGroup.PLANT_BASED_CORE,
                volume_kg=Decimal("100"),
                protein_pct=Decimal("20"),
                protein_kg=Decimal("20"),
                used_per_product_split=False,
                methodology_version="1.0.0",
                methodology_source_edition="demo",
                taxonomy_version="1.0.0",
                rules_version="0.1.0",
            )
            for p in products
        ]
        store.add_run(
            RunRecord(
                id=run_id,
                project_id=UUID(project_id),
                methodology=Methodology.PROTEIN_TRACKER,
                started_at=now,
                finished_at=now,
                triggered_by=store.default_user_id,
                rows_payload=[r.model_dump() for r in rows],
                summary_payload={},
                rows_count=len(rows),
                organisation_id=store.default_org_id,
            )
        )

        r = client.get(f"/api/v1/projects/{project_id}/export/categorized.xlsx")
        assert r.status_code == 200, r.text
        ws = load_workbook(BytesIO(r.content))["Produits"]
        hdr = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        assert "Protéines végétales (kg)" in hdr
        plant_col = hdr.index("Protéines végétales (kg)") + 1
        animal_col = hdr.index("Protéines animales (kg)") + 1
        # all-plant (PLANT_BASED_CORE) → 20 kg plant, 0 animal.
        assert ws.cell(row=2, column=plant_col).value == 20
        assert ws.cell(row=2, column=animal_col).value == 0
